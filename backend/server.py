from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import httpx
from enum import Enum
import base64
from io import BytesIO
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from fastapi.responses import StreamingResponse
import csv
from urllib.parse import unquote, urlparse, parse_qs
import aiohttp
import subprocess
import sys

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'sharkdrop_wms')]

# Helper to clean MongoDB documents
def clean_doc(doc):
    """Remove MongoDB _id field from document"""
    if doc and '_id' in doc:
        del doc['_id']
    return doc

def clean_docs(docs):
    """Remove MongoDB _id field from list of documents"""
    return [clean_doc(doc) for doc in docs]

# JWT Settings
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'sharkdrop-super-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()
security_optional = HTTPBearer(auto_error=False)

# Create the main app
app = FastAPI(title="SharkDrop WMS API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Shopify sync control (single-process guard)
shopify_sync_lock = asyncio.Lock()
shopify_sync_task = None

# Shopify variant UPC backup metafield
UPC_BACKUP_NAMESPACE = os.getenv("UPC_BACKUP_NAMESPACE", "custom").strip() or "custom"
UPC_BACKUP_KEY = os.getenv("UPC_BACKUP_KEY", "upc_backup").strip() or "upc_backup"

def get_shopify_config():
    shop_domain = os.environ.get("SHOPIFY_SHOP_DOMAIN") or os.environ.get("SHOP_DOMAIN")
    access_token = os.environ.get("SHOPIFY_ACCESS_TOKEN") or os.environ.get("ADMIN_API_TOKEN")
    api_version = os.environ.get("SHOPIFY_API_VERSION") or os.environ.get("API_VERSION") or "2024-07"
    return shop_domain, access_token, api_version

def _parse_ids(value: Optional[str]) -> Optional[List[str]]:
    if not value:
        return None
    return [v.strip() for v in value.split(",") if v.strip()]

def _normalize_barcode(value: str) -> str:
    raw = (value or "").strip()
    digits = re.sub(r"\D+", "", raw)
    return digits if digits else raw

async def _kicks_request(session: aiohttp.ClientSession, url: str, params: dict) -> dict:
    try:
        async with session.get(url, params=params, timeout=45) as resp:
            if resp.status != 200:
                return {}
            return await resp.json()
    except Exception:
        return {}

def _extract_stockx_product(data: Any) -> Optional[dict]:
    if not data:
        return None
    if isinstance(data, dict) and "data" in data:
        data = data.get("data")
    if isinstance(data, dict) and "products" in data and isinstance(data.get("products"), list):
        data = data.get("products")
    if isinstance(data, dict) and "product" in data and isinstance(data.get("product"), dict):
        data = data.get("product")
    if isinstance(data, list):
        return data[0] if data else None
    if isinstance(data, dict):
        return data
    return None

def _extract_products_list(data: Any) -> List[dict]:
    if not data:
        return []
    if isinstance(data, dict) and "data" in data:
        data = data.get("data")
    if isinstance(data, dict) and "products" in data and isinstance(data.get("products"), list):
        return data.get("products") or []
    if isinstance(data, dict) and "results" in data and isinstance(data.get("results"), list):
        return data.get("results") or []
    if isinstance(data, dict) and "gtins" in data and isinstance(data.get("gtins"), list):
        return data.get("gtins") or []
    if isinstance(data, dict) and "items" in data and isinstance(data.get("items"), list):
        return data.get("items") or []
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    return []

def _barcode_candidates(value: str) -> List[str]:
    raw = (value or "").strip()
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return []
    out: List[str] = []
    def add(x: str):
        v = re.sub(r"\D+", "", (x or "").strip())
        if not v:
            return
        if v not in out:
            out.append(v)

    # Base value.
    add(digits)

    # UPC-12 -> EAN-13
    if len(digits) == 12:
        add("0" + digits)

    # EAN-13 with leading 0 -> UPC-12
    if len(digits) == 13 and digits.startswith("0"):
        add(digits[1:])

    # GTIN-14 common conversions.
    if len(digits) == 14 and digits.startswith("0"):
        d13 = digits[1:]
        add(d13)
        if len(d13) == 13 and d13.startswith("0"):
            add(d13[1:])

    return out

def _product_matches_barcode(full_product: dict, candidates: List[str]) -> bool:
    if not full_product or not candidates:
        return False
    cand = set(candidates)
    # Check generic identifier containers first
    for ident in (full_product.get("identifiers") or []):
        val = re.sub(r"\D+", "", (ident.get("identifier") or ""))
        if not val:
            continue
        if val in cand:
            return True
        if len(val) == 12 and ("0" + val) in cand:
            return True
        if len(val) in (13, 14) and val.startswith("0") and val[1:] in cand:
            return True
    for v in (full_product.get("variants") or []):
        for ident in (v.get("identifiers") or []):
            val = re.sub(r"\D+", "", (ident.get("identifier") or ""))
            if not val:
                continue
            if val in cand:
                return True
            if len(val) == 12 and ("0" + val) in cand:
                return True
            if len(val) in (13, 14) and val.startswith("0") and val[1:] in cand:
                return True
    return False


async def _upsert_stockx_overrides_from_match(barcode: str, match: dict) -> None:
    identifiers = set()
    for cand in _barcode_candidates(barcode):
        identifiers.add(cand)
    identifiers.add(_normalize_barcode(barcode))

    slug = (match or {}).get("slug")
    if slug:
        full = await _kicks_get_product_full(slug)
        for v in (full or {}).get("variants", []):
            for ident in (v.get("identifiers") or []):
                val = re.sub(r"\D+", "", (ident.get("identifier") or ""))
                if val and len(val) in (12, 13, 14):
                    identifiers.add(val)

    for ident in list(identifiers)[:80]:
        if not ident:
            continue
        await db.stockx_barcode_overrides.update_one(
            {"normalized_barcode": ident},
            {
                "$set": {
                    "barcode": ident,
                    "normalized_barcode": ident,
                    "match": match,
                    "updated_at": datetime.utcnow(),
                }
            },
            upsert=True,
        )

def _stockx_url_from_product(product: dict) -> Optional[str]:
    url = (product.get("url") or product.get("productUrl") or "").strip()
    if url and "stockx.com" in url:
        return url
    slug = (product.get("slug") or product.get("url_slug") or product.get("urlSlug") or "").strip()
    if slug:
        return f"https://stockx.com/{slug}"
    return None

def _stockx_slug_from_url(url: str) -> Optional[str]:
    if not url:
        return None
    m = re.search(r"stockx\.com/([^?\s#]+)", url.strip(), flags=re.I)
    if not m:
        return None
    slug = m.group(1).strip().strip("/")
    return slug or None

def _parse_size_number(raw: str) -> Optional[float]:
    s = (raw or "").strip().upper()
    if not s:
        return None
    s = s.replace(",", ".")
    s = s.replace("Â½", ".5")
    s = s.replace("⅓", ".333").replace("⅔", ".666")
    s = s.replace("1/2", ".5").replace("1/3", ".333").replace("2/3", ".666")
    s = re.sub(r"\s+", " ", s)
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except:
        return None

USW_TO_EU = {
    5.0: 35.5, 5.5: 36.0, 6.0: 36.5, 6.5: 37.5, 7.0: 38.0, 7.5: 38.5,
    8.0: 39.0, 8.5: 40.0, 9.0: 40.5, 9.5: 41.0, 10.0: 42.0, 10.5: 42.5,
    11.0: 43.0, 11.5: 44.0, 12.0: 44.5
}
USM_TO_EU = {
    4.0: 36.0, 4.5: 36.5, 5.0: 37.5, 5.5: 38.0, 6.0: 38.5, 6.5: 39.0,
    7.0: 40.0, 7.5: 40.5, 8.0: 41.0, 8.5: 42.0, 9.0: 42.5, 9.5: 43.0,
    10.0: 44.0, 10.5: 44.5, 11.0: 45.0, 11.5: 45.5, 12.0: 46.0
}

def _approx_lookup(mapping: dict, us: float) -> Optional[float]:
    if us in mapping:
        return mapping[us]
    rounded = round(us * 2) / 2
    return mapping.get(rounded)

def _get_eu_size_from_variant(v: dict, title_hint: str, slug_hint: str) -> Optional[float]:
    sizes = (v.get("sizes") or []) if isinstance(v, dict) else []
    for s in sizes:
        t = (s.get("type") or "").strip().lower()
        if t == "eu":
            raw = (s.get("size") or "").strip()
            val = _parse_size_number(raw)
            if val is not None:
                return val

    womens = ("WOMEN" in (title_hint or "").upper()) or ("-womens" in slug_hint.lower()) or slug_hint.lower().endswith("womens")
    candidates = []
    for s in sizes:
        t = (s.get("type") or "").strip().lower().replace(" ", "")
        raw = (s.get("size") or "").strip()
        if "us" in t:
            n = _parse_size_number(raw)
            if n is not None:
                candidates.append((t, n))

    if not candidates:
        return None

    usw = None
    usm = None
    for (t, n) in candidates:
        if "w" in t or "women" in t:
            usw = n
            break
    if usw is None:
        if womens:
            usw = candidates[0][1]
        else:
            usm = candidates[0][1]
    if usw is not None:
        return _approx_lookup(USW_TO_EU, usw)
    if usm is not None:
        return _approx_lookup(USM_TO_EU, usm)
    return None

def _eu_display(eu: float) -> str:
    if eu is None:
        return ""
    if abs(eu - round(eu)) < 1e-9:
        return str(int(round(eu)))
    if abs(eu - (int(eu) + 0.5)) < 1e-9:
        return f"{int(eu)}.5"
    return f"{eu:.3f}".rstrip("0").rstrip(".")

def _variant_size_numeric(title: str) -> Optional[float]:
    m = re.search(r"(\d+(?:\.\d+)?)", (title or "").replace(",", "."))
    if not m:
        return None
    try:
        return float(m.group(1))
    except:
        return None

async def _kicks_get_product_full(slug: str) -> dict:
    key = os.environ.get("KICKS_API_KEY")
    if not key:
        return {}
    market = (os.environ.get("KICKS_MARKET") or "IT").upper()
    headers = {"Authorization": f"Bearer {key}"}
    url = f"https://api.kicks.dev/v3/stockx/products/{slug}"
    params = {
        "display[variants]": "true",
        "display[prices]": "true",
        "display[identifiers]": "true",
        "display[gallery]": "true",
        "display[traits]": "true",
        "market": market,
    }
    async with aiohttp.ClientSession(headers=headers) as session:
        data = await _kicks_request(session, url, params)
    return (data or {}).get("data") or {}

async def _add_unpriced_variants_to_wms(shopify_product_id: str, stockx_url: str, express_mode: str, express_label: str):
    slug = _stockx_slug_from_url(stockx_url)
    if not slug:
        return 0

    product = None
    for _ in range(10):
        product = await db.products.find_one({"shopify_product_id": str(shopify_product_id)}, {"_id": 0})
        if product:
            break
        await asyncio.sleep(10)
    if not product:
        return 0

    kicks = await _kicks_get_product_full(slug)
    if not kicks:
        return 0

    title = (kicks.get("title") or slug).strip()
    sku_base = (kicks.get("sku") or slug.upper()).strip()

    existing_sizes = set()
    for v in product.get("variants", []):
        n = _variant_size_numeric(v.get("title") or "")
        if n is not None:
            existing_sizes.add(round(n * 2) / 2)

    new_variants = []
    for v in (kicks.get("variants") or []):
        eu_val = _get_eu_size_from_variant(v, title_hint=title, slug_hint=slug)
        if eu_val is None:
            continue
        size_key = round(eu_val * 2) / 2
        if size_key in existing_sizes:
            continue

        eu_disp = _eu_display(eu_val)
        if not eu_disp:
            continue
        option1 = eu_disp
        if express_mode == "all":
            option1 = f"{eu_disp} - {express_label}"

        barcode = None
        upc_backup = None
        identifiers = v.get("identifiers") or []
        for x in identifiers:
            t = (x.get("identifier_type") or x.get("identifierType") or "").strip().upper()
            val = re.sub(r"\D+", "", (x.get("identifier") or ""))
            if not val:
                continue
            if len(val) == 13 and barcode is None:
                barcode = val
            if len(val) == 14 and barcode is None:
                barcode = val
            if len(val) == 12 and upc_backup is None:
                upc_backup = val
        if not barcode and upc_backup:
            barcode = "0" + upc_backup

        new_variants.append({
            "id": str(uuid.uuid4()),
            "title": option1,
            "sku": sku_base,
            "barcode": barcode,
            "upc_backup": upc_backup,
            "price": 0,
            "shopify_variant_id": None,
            "inventory_item_id": None,
            "inventory_management": None,
            "inventory_quantity": 0,
            "source": "stockx_unpriced",
        })

    # Update existing variants with UPC backup if missing (match by size)
    existing_updates = 0
    if product.get("variants"):
        updates = []
        for vv in product.get("variants", []):
            size_val = _variant_size_numeric(vv.get("title") or "")
            if size_val is None:
                continue
            size_key = round(size_val * 2) / 2
            for nv in new_variants:
                n_size = _variant_size_numeric(nv.get("title") or "")
                if n_size is None:
                    continue
                if round(n_size * 2) / 2 != size_key:
                    continue
                if vv.get("upc_backup"):
                    break
                vv["upc_backup"] = nv.get("upc_backup")
                if not vv.get("barcode") and nv.get("barcode"):
                    vv["barcode"] = nv.get("barcode")
                updates.append(vv)
                break
        if updates:
            existing_updates = len(updates)
            await db.products.update_one(
                {"shopify_product_id": str(shopify_product_id)},
                {"$set": {"variants": product.get("variants", []), "updated_at": datetime.utcnow()}}
            )

    if not new_variants and existing_updates == 0:
        return 0

    if new_variants:
        await db.products.update_one(
            {"shopify_product_id": str(shopify_product_id)},
            {"$push": {"variants": {"$each": new_variants}}, "$set": {"updated_at": datetime.utcnow()}}
        )
    return len(new_variants)

async def kicks_lookup_by_barcode(barcode: str) -> Optional[dict]:
    key = os.environ.get("KICKS_API_KEY")
    if not key:
        return None
    market = (os.environ.get("KICKS_MARKET") or "IT").upper()
    headers = {"Authorization": f"Bearer {key}"}
    base = "https://api.kicks.dev/v3/stockx"
    candidates = _barcode_candidates(barcode)
    if not candidates:
        return {"error": "unsupported_barcode"}
    async with aiohttp.ClientSession(headers=headers) as session:
        for candidate in candidates:
            # TROVALINK strategy: explicit StockX barcodes filter is the most reliable lookup.
            queries = [
                ("products", {"filters": f'barcodes = "{candidate}"', "limit": 25, "market": market}),
                ("products", {"query": candidate, "limit": 25, "market": market, "display[identifiers]": "true"}),
                ("products", {"filter[identifiers]": candidate, "market": market}),
                ("products", {"filter[identifier]": candidate, "market": market}),
                ("products", {"filter[barcode]": candidate, "market": market}),
                ("products", {"filter[gtin]": candidate, "market": market}),
                ("products", {"filter[gtins]": candidate, "market": market}),
                ("products", {"search": candidate, "market": market}),
                ("search", {"query": candidate, "market": market}),
            ]
            for path, params in queries:
                data = await _kicks_request(session, f"{base}/{path}", params)
                products = _extract_products_list(data)
                for product in products:
                    product_barcodes = [re.sub(r"\D+", "", str(x or "")) for x in (product.get("barcodes") or [])]
                    has_barcode_match = any(b in set(candidates) for b in product_barcodes if b)
                    slug = (product.get("slug") or product.get("url_slug") or product.get("urlSlug") or "").strip()
                    if not slug:
                        continue
                    if has_barcode_match:
                        stockx_url = _stockx_url_from_product(product) or f"https://stockx.com/{slug}"
                        return {
                            "title": product.get("title") or product.get("name"),
                            "brand": product.get("brand") or product.get("vendor"),
                            "slug": slug,
                            "image_url": product.get("image") or product.get("imageUrl") or (product.get("images") or [None])[0],
                            "stockx_url": stockx_url,
                            "verified": True,
                            "raw": product,
                            "verified_source": "v3_barcodes",
                        }
                    full = await _kicks_get_product_full(slug)
                    if not full:
                        continue
                    if not _product_matches_barcode(full, candidates):
                        continue
                    stockx_url = _stockx_url_from_product(product) or _stockx_url_from_product(full)
                    return {
                        "title": full.get("title") or product.get("title") or product.get("name"),
                        "brand": full.get("brand") or product.get("brand") or product.get("vendor"),
                        "slug": slug,
                        "image_url": full.get("image") or product.get("image") or product.get("imageUrl") or (product.get("images") or [None])[0],
                        "stockx_url": stockx_url,
                        "verified": True,
                        "raw": product,
                    }
    # Fallback: v2 GTIN endpoint (more reliable for barcodes)
    async with aiohttp.ClientSession(headers={"Authorization": f"Bearer {key}"}) as session:
        for candidate in candidates:
            if len(candidate) == 12:
                btype = "UPC"
            elif len(candidate) == 13:
                btype = "EAN-13"
            elif len(candidate) == 14:
                btype = "ITF-14"
            else:
                btype = None

            params = {"barcode": candidate}
            if btype:
                params["barcode_type"] = btype

            data = await _kicks_request(session, "https://api.kicks.dev/v2/gtins", params)
            items = _extract_products_list(data)
            if not items:
                continue

            item = items[0]
            link = (item.get("link") or "").strip()
            stockx_url = None
            if link:
                parsed = urlparse(link)
                q = parse_qs(parsed.query or "")
                if "u" in q and q["u"]:
                    stockx_url = unquote(q["u"][0])
                else:
                    stockx_url = link

            slug = _stockx_slug_from_url(stockx_url or "")
            full = await _kicks_get_product_full(slug) if slug else {}

            return {
                "title": (full.get("title") if full else None) or item.get("title"),
                "brand": (full.get("brand") if full else None) or item.get("brand"),
                "slug": slug,
                "image_url": (full.get("image") if full else None) or item.get("image"),
                "stockx_url": stockx_url,
                "verified": True,
                "raw": item,
                "verified_source": "gtins",
            }

    # Last resort: try v3 gtins endpoint if available
    async with aiohttp.ClientSession(headers={"Authorization": f"Bearer {key}"}) as session:
        for candidate in candidates:
            params = {"barcode": candidate, "market": market}
            data = await _kicks_request(session, "https://api.kicks.dev/v3/stockx/gtins", params)
            items = _extract_products_list(data)
            if not items:
                continue

            item = items[0]
            stockx_url = _stockx_url_from_product(item)
            slug = _stockx_slug_from_url(stockx_url or "")
            full = await _kicks_get_product_full(slug) if slug else {}

            return {
                "title": (full.get("title") if full else None) or item.get("title"),
                "brand": (full.get("brand") if full else None) or item.get("brand"),
                "slug": slug,
                "image_url": (full.get("image") if full else None) or item.get("image"),
                "stockx_url": stockx_url,
                "verified": True,
                "raw": item,
                "verified_source": "gtins_v3",
            }

    return None

async def kicks_lookup_by_query(query: str) -> Optional[dict]:
    key = os.environ.get("KICKS_API_KEY")
    if not key:
        return None
    q = (query or "").strip()
    if not q:
        return None
    market = (os.environ.get("KICKS_MARKET") or "IT").upper()
    headers = {"Authorization": f"Bearer {key}"}
    base = "https://api.kicks.dev/v3/stockx"
    async with aiohttp.ClientSession(headers=headers) as session:
        queries = [
            ("products", {"search": q, "limit": 25, "market": market}),
            ("products", {"query": q, "limit": 25, "market": market}),
            ("search", {"query": q, "limit": 25, "market": market}),
        ]
        for path, params in queries:
            data = await _kicks_request(session, f"{base}/{path}", params)
            products = _extract_products_list(data)
            if not products:
                continue

            q_cf = q.casefold()
            exact = [p for p in products if (p.get("title") or p.get("name") or "").casefold() == q_cf]
            picked = exact[0] if exact else None
            if not picked:
                products.sort(key=lambda p: (-(p.get("rank") or 0)))
                picked = products[0]

            slug = (picked.get("slug") or picked.get("url_slug") or picked.get("urlSlug") or "").strip()
            if not slug:
                continue

            full = await _kicks_get_product_full(slug)
            source = full or picked
            stockx_url = _stockx_url_from_product(source) or f"https://stockx.com/{slug}"
            return {
                "title": source.get("title") or source.get("name"),
                "brand": source.get("brand") or source.get("vendor"),
                "slug": slug,
                "image_url": source.get("image") or source.get("imageUrl") or (source.get("images") or [None])[0],
                "stockx_url": stockx_url,
                "verified": False,
                "raw": picked,
                "verified_source": "query",
            }
    return None

# ==================== ENUMS ====================
class TransactionType(str, Enum):
    RECEIVE = "receive"
    MOVE = "move"
    TRANSFER = "transfer"
    SALE = "sale"
    ADJUST = "adjust"

class UserRole(str, Enum):
    ADMIN = "admin"
    OPERATOR = "operator"

class ActionType(str, Enum):
    # Auth actions
    LOGIN = "login"
    LOGOUT = "logout"
    # User management
    USER_CREATE = "user_create"
    USER_UPDATE = "user_update"
    USER_BLOCK = "user_block"
    USER_UNBLOCK = "user_unblock"
    # Location/Shelf actions
    LOCATION_CREATE = "location_create"
    LOCATION_UPDATE = "location_update"
    LOCATION_DELETE = "location_delete"
    SHELF_CREATE = "shelf_create"
    SHELF_UPDATE = "shelf_update"
    SHELF_DELETE = "shelf_delete"
    # Inventory actions
    INVENTORY_RECEIVE = "inventory_receive"
    INVENTORY_MOVE = "inventory_move"
    INVENTORY_TRANSFER = "inventory_transfer"
    INVENTORY_SALE = "inventory_sale"
    INVENTORY_ADJUST = "inventory_adjust"
    STALL_CREATE = "stall_create"
    STALL_RETURN = "stall_return"
    STALL_SELL = "stall_sell"
    STALL_MOVE = "stall_move"
    # Rollback
    ROLLBACK = "rollback"
    # Shopify
    SHOPIFY_SYNC = "shopify_sync"
    SHOPIFY_UPDATE = "shopify_update"

# ==================== MODELS ====================

# User Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    role: UserRole = UserRole.OPERATOR
    created_at: datetime = Field(default_factory=datetime.utcnow)
    created_by: Optional[str] = None  # admin who created this user
    is_active: bool = True
    blocked_at: Optional[datetime] = None
    blocked_by: Optional[str] = None
    blocked_reason: Optional[str] = None

class UserCreate(BaseModel):
    username: str
    password: str
    role: UserRole = UserRole.OPERATOR

class UserUpdate(BaseModel):
    password: Optional[str] = None
    role: Optional[UserRole] = None

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    role: UserRole
    is_active: bool

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# Location Models
class Location(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    shopify_location_id: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True

class LocationCreate(BaseModel):
    name: str
    description: Optional[str] = None
    shopify_location_id: Optional[str] = None

# Shelf Models
class Shelf(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # e.g., A1, A2, B1
    barcode: str  # e.g., SD-0000A1
    location_id: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True

class ShelfCreate(BaseModel):
    name: str
    barcode: str
    location_id: str

# Product Models (cached from Shopify)
class ProductVariant(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shopify_variant_id: Optional[str] = None
    product_id: str
    title: str  # e.g., "S", "M", "L"
    sku: Optional[str] = None
    barcode: Optional[str] = None
    upc_backup: Optional[str] = None
    price: Optional[float] = None
    inventory_item_id: Optional[str] = None
    inventory_management: Optional[str] = None
    inventory_quantity: Optional[int] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class Product(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shopify_product_id: Optional[str] = None
    title: str
    handle: Optional[str] = None
    image_url: Optional[str] = None
    image_base64: Optional[str] = None  # Store image as base64
    tags: List[str] = []
    variants: List[ProductVariant] = []
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True

class ProductCreate(BaseModel):
    title: str
    image_url: Optional[str] = None
    variants: List[dict] = []

# Inventory Level Models
class InventoryLevel(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    variant_id: str
    product_id: Optional[str] = None
    location_id: str
    shelf_id: Optional[str] = None
    quantity: int = 0
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    source: Optional[str] = "local"  # local | shopify
    last_synced_at: Optional[datetime] = None

# Inventory Transaction Models (immutable log)
class InventoryTransaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_type: TransactionType
    variant_id: str
    product_id: str
    quantity: int
    from_location_id: Optional[str] = None
    to_location_id: Optional[str] = None
    from_shelf_id: Optional[str] = None
    to_shelf_id: Optional[str] = None
    sale_price: Optional[float] = None
    note: Optional[str] = None
    user_id: str
    username: str = ""  # Store username for easy display
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_rolled_back: bool = False
    rolled_back_at: Optional[datetime] = None
    rolled_back_by: Optional[str] = None

# Stall (temporary out of shelf)
class StallItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    variant_id: str
    product_id: str
    quantity: int
    from_location_id: str
    from_shelf_id: Optional[str] = None
    note: Optional[str] = None
    customer_name: Optional[str] = None
    status: str = "in_stallo"  # in_stallo | returned | sold | moved
    resolved_action: Optional[str] = None
    to_location_id: Optional[str] = None
    to_shelf_id: Optional[str] = None
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    resolved_at: Optional[datetime] = None

class StallCreateRequest(BaseModel):
    variant_id: str
    location_id: str
    shelf_id: Optional[str] = None
    quantity: int = 1
    note: Optional[str] = None
    customer_name: Optional[str] = None

class StallReturnRequest(BaseModel):
    to_location_id: Optional[str] = None
    to_shelf_id: Optional[str] = None

class StallSellRequest(BaseModel):
    sale_price: Optional[float] = None

class StallMoveRequest(BaseModel):
    to_location_id: str
    to_shelf_id: Optional[str] = None

# Action Log Models (for complete audit trail)
class ActionLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    action_type: ActionType
    user_id: str
    username: str
    description: str
    entity_type: Optional[str] = None  # product, location, shelf, inventory, user
    entity_id: Optional[str] = None
    old_data: Optional[Dict[str, Any]] = None  # snapshot before change
    new_data: Optional[Dict[str, Any]] = None  # snapshot after change
    metadata: Optional[Dict[str, Any]] = None  # additional info
    created_at: datetime = Field(default_factory=datetime.utcnow)
    ip_address: Optional[str] = None

class SystemLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    level: str
    message: str
    context: Optional[Dict[str, Any]] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ReceiveRequest(BaseModel):
    variant_id: str
    location_id: str
    shelf_id: Optional[str] = None
    quantity: int

class MoveRequest(BaseModel):
    variant_id: str
    location_id: str
    from_shelf_id: str
    to_shelf_id: str
    quantity: int

class TransferRequest(BaseModel):
    variant_id: str
    from_location_id: str
    to_location_id: str
    from_shelf_id: Optional[str] = None
    to_shelf_id: Optional[str] = None
    quantity: int

class SaleRequest(BaseModel):
    variant_id: str
    location_id: str
    shelf_id: Optional[str] = None
    quantity: int
    sale_price: float

class AdjustRequest(BaseModel):
    variant_id: str
    location_id: str
    shelf_id: Optional[str] = None
    new_quantity: int
    note: Optional[str] = None

# Shopify Sync State
class ShopifySyncState(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    last_sync_at: Optional[datetime] = None
    products_synced: int = 0
    status: str = "idle"  # idle, syncing, completed, error
    error_message: Optional[str] = None

# Collection Model (from Shopify)
class Collection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shopify_collection_id: Optional[str] = None
    title: str
    handle: Optional[str] = None
    product_ids: List[str] = []  # List of product IDs in this collection
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True

class CollectionCreate(BaseModel):
    title: str
    product_ids: List[str] = []

# Update models for editing
class LocationUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    shopify_location_id: Optional[str] = None

class ShelfUpdate(BaseModel):
    name: Optional[str] = None
    barcode: Optional[str] = None
    location_id: Optional[str] = None

class ProductUpdate(BaseModel):
    title: Optional[str] = None
    image_url: Optional[str] = None

# StockX lookup/import
class StockxLookupRequest(BaseModel):
    barcode: str

class StockxLookupUrlRequest(BaseModel):
    url: str
    barcode: Optional[str] = None

class StockxLookupQueryRequest(BaseModel):
    query: str
    barcode: Optional[str] = None

class StockxImportRequest(BaseModel):
    barcode: str
    stockx_url: str
    product_type: str = "Scarpe"
    price_mode: str = "fixed"  # fixed|market
    fixed_price_eur: Optional[float] = None
    price_type_preferred: Optional[str] = "standard"
    size_mode: str = "all"  # all|range|list
    eu_min: Optional[float] = None
    eu_max: Optional[float] = None
    eu_list: Optional[List[float]] = None
    express_mode: str = "none"  # all|none
    express_label: Optional[str] = "Express 24/48H"
    tags: Optional[str] = None
    template_suffix: Optional[str] = None
    status: Optional[str] = "active"
    default_qty: Optional[int] = 0
    lookup_id: Optional[str] = None

class VariantUpdate(BaseModel):
    title: Optional[str] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None
    upc_backup: Optional[str] = None
    price: Optional[float] = None

# ==================== LOCAL PRODUCT MODELS ====================

class LocalVariantCreate(BaseModel):
    title: str  # e.g., "S", "M", "L", "42", "43"
    sku: Optional[str] = None
    barcode: Optional[str] = None
    price: float

class LocalProductCreate(BaseModel):
    title: str
    description: Optional[str] = None
    image_base64: Optional[str] = None
    price: float  # Default price if no variants
    sku: Optional[str] = None
    barcode: Optional[str] = None
    variants: List[LocalVariantCreate] = []  # Empty = single product, non-empty = with sizes

# ==================== PURCHASE FROM SUPPLIER MODELS ====================

class PurchaseItemCreate(BaseModel):
    product_id: Optional[str] = None
    variant_id: Optional[str] = None
    title: str
    variant_title: Optional[str] = None
    quantity: int = 1
    purchase_price: float = 0.0

class PurchaseLinkCreate(BaseModel):
    items: List[PurchaseItemCreate]
    note: Optional[str] = None
    doc_type: Optional[str] = "acquisto"  # acquisto|contovendita

class SupplierData(BaseModel):
    first_name: str
    last_name: str
    birth_date: str  # DD/MM/YYYY
    birth_place: str
    birth_country: str
    residence_address: str
    residence_city: str
    residence_province: str
    residence_cap: str
    residence_country: str
    fiscal_code: str
    iban: Optional[str] = None
    signature: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class PurchaseLink(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    token: str = Field(default_factory=lambda: str(uuid.uuid4()).replace("-", ""))
    items: List[dict]  # List of purchase items
    total_amount: float
    note: Optional[str] = None
    created_by: str
    created_by_username: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    expires_at: datetime  # 2 days from creation
    status: str = "pending"  # pending, submitted, completed, expired
    supplier_data: Optional[dict] = None
    submitted_at: Optional[datetime] = None
    doc_type: str = "acquisto"  # acquisto|contovendita

# ==================== AUTH HELPERS ====================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Invalid credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"id": user_id})
    if user is None:
        raise credentials_exception
    return user

async def _get_user_from_token(token: str) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Invalid credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = await db.users.find_one({"id": user_id})
    if user is None:
        raise credentials_exception
    return user

async def get_current_user_or_token(
    token: Optional[str] = None,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_optional)
) -> dict:
    if token:
        return await _get_user_from_token(token)
    if credentials and credentials.credentials:
        return await _get_user_from_token(credentials.credentials)
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Not authenticated")

async def require_admin(current_user: dict = Depends(get_current_user)) -> dict:
    if current_user.get("role") != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user

# ==================== ACTION LOG HELPER ====================

async def log_action(
    action_type: ActionType,
    user: dict,
    description: str,
    entity_type: str = None,
    entity_id: str = None,
    old_data: dict = None,
    new_data: dict = None,
    metadata: dict = None
):
    """Log an action for audit trail"""
    log_entry = ActionLog(
        action_type=action_type,
        user_id=user.get("id", "system"),
        username=user.get("username", "system"),
        description=description,
        entity_type=entity_type,
        entity_id=entity_id,
        old_data=old_data,
        new_data=new_data,
        metadata=metadata
    )
    await db.action_logs.insert_one(log_entry.dict())
    logger.info(f"ACTION: {action_type} by {user.get('username')} - {description}")

async def log_system_event(level: str, message: str, context: dict = None):
    """Persist important system events for diagnostics."""
    entry = SystemLog(level=level, message=message, context=context)
    await db.system_logs.insert_one(entry.dict())
    if level.lower() in ("error", "critical"):
        logger.error(f"SYSTEM[{level}]: {message}")
    else:
        logger.info(f"SYSTEM[{level}]: {message}")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    # Check if username exists
    existing = await db.users.find_one({"username": user_data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Create user
    user = User(
        username=user_data.username,
        password_hash=get_password_hash(user_data.password),
        role=user_data.role
    )
    await db.users.insert_one(user.dict())
    
    # Create token
    access_token = create_access_token(data={"sub": user.id})
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(id=user.id, username=user.username, role=user.role, is_active=user.is_active)
    )

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"username": user_data.username})
    if not user or not verify_password(user_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="User is inactive")
    
    access_token = create_access_token(data={"sub": user["id"]})
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(id=user["id"], username=user["username"], role=user["role"], is_active=user.get("is_active", True))
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        username=current_user["username"],
        role=current_user["role"],
        is_active=current_user.get("is_active", True)
    )

# ==================== USER MANAGEMENT ROUTES (Admin Only) ====================

@api_router.get("/users")
async def get_users(current_user: dict = Depends(require_admin)):
    """Get all users (admin only)"""
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.post("/users")
async def create_user(data: UserCreate, current_user: dict = Depends(require_admin)):
    """Create a new user (admin only)"""
    # Check if username exists
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Create user
    user = User(
        username=data.username,
        password_hash=get_password_hash(data.password),
        role=data.role,
        created_by=current_user["id"]
    )
    await db.users.insert_one(user.dict())
    
    # Log action
    await log_action(
        ActionType.USER_CREATE,
        current_user,
        f"Created user '{data.username}' with role {data.role}",
        entity_type="user",
        entity_id=user.id,
        new_data={"username": data.username, "role": data.role}
    )
    
    return {"id": user.id, "username": user.username, "role": user.role, "is_active": user.is_active}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, current_user: dict = Depends(require_admin)):
    """Update a user (admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {}
    old_data = {}
    
    if data.password:
        update_data["password_hash"] = get_password_hash(data.password)
    if data.role:
        old_data["role"] = user.get("role")
        update_data["role"] = data.role
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        
        await log_action(
            ActionType.USER_UPDATE,
            current_user,
            f"Updated user '{user['username']}'",
            entity_type="user",
            entity_id=user_id,
            old_data=old_data,
            new_data={"role": data.role} if data.role else {}
        )
    
    return {"message": "User updated"}

@api_router.post("/users/{user_id}/block")
async def block_user(user_id: str, reason: str = None, current_user: dict = Depends(require_admin)):
    """Block a user (admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user["id"] == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot block yourself")
    
    if user.get("role") == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Cannot block admin users")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "is_active": False,
            "blocked_at": datetime.utcnow(),
            "blocked_by": current_user["id"],
            "blocked_reason": reason
        }}
    )
    
    await log_action(
        ActionType.USER_BLOCK,
        current_user,
        f"Blocked user '{user['username']}'" + (f" - Reason: {reason}" if reason else ""),
        entity_type="user",
        entity_id=user_id,
        metadata={"reason": reason}
    )
    
    return {"message": f"User '{user['username']}' blocked"}

@api_router.post("/users/{user_id}/unblock")
async def unblock_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Unblock a user (admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_active": True}, "$unset": {"blocked_at": "", "blocked_by": "", "blocked_reason": ""}}
    )
    
    await log_action(
        ActionType.USER_UNBLOCK,
        current_user,
        f"Unblocked user '{user['username']}'",
        entity_type="user",
        entity_id=user_id
    )
    
    return {"message": f"User '{user['username']}' unblocked"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_admin)):
    """Delete a user (admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user["id"] == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    await db.users.delete_one({"id": user_id})
    
    await log_action(
        ActionType.USER_CREATE,
        current_user,
        f"Deleted user '{user['username']}'",
        entity_type="user",
        entity_id=user_id
    )
    
    return {"message": "User deleted"}

# ==================== ACTION LOGS ROUTES (Admin Only) ====================

@api_router.get("/logs")
async def get_action_logs(
    skip: int = 0,
    limit: int = 50,
    user_id: Optional[str] = None,
    action_type: Optional[str] = None,
    entity_type: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    """Get action logs (admin only)"""
    query = {}
    if user_id:
        query["user_id"] = user_id
    if action_type:
        query["action_type"] = action_type
    if entity_type:
        query["entity_type"] = entity_type
    
    total = await db.action_logs.count_documents(query)
    logs = await db.action_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {"total": total, "logs": logs}

@api_router.get("/logs/user/{user_id}")
async def get_user_logs(user_id: str, skip: int = 0, limit: int = 50, current_user: dict = Depends(require_admin)):
    """Get logs for a specific user (admin only)"""
    total = await db.action_logs.count_documents({"user_id": user_id})
    logs = await db.action_logs.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"total": total, "logs": logs}

@api_router.get("/system-logs")
async def get_system_logs(
    level: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: dict = Depends(require_admin)
):
    """Get system logs (admin only)"""
    query = {}
    if level:
        query["level"] = level
    total = await db.system_logs.count_documents(query)
    logs = await db.system_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(offset).limit(limit).to_list(limit)
    return {"total": total, "logs": logs, "limit": limit, "offset": offset}

@api_router.get("/transactions")
async def get_inventory_transactions(
    skip: int = 0,
    limit: int = 50,
    user_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    include_rolled_back: bool = False,
    current_user: dict = Depends(require_admin)
):
    """Get inventory transactions (admin only)"""
    query = {}
    if user_id:
        query["user_id"] = user_id
    if transaction_type:
        query["transaction_type"] = transaction_type
    if not include_rolled_back:
        query["is_rolled_back"] = {"$ne": True}
    
    total = await db.inventory_transactions.count_documents(query)
    transactions = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with product info
    for tx in transactions:
        product = await db.products.find_one({"id": tx.get("product_id")}, {"_id": 0, "title": 1, "variants": 1})
        if product:
            tx["product_title"] = product.get("title")
            # Find variant title
            for v in product.get("variants", []):
                if v.get("id") == tx.get("variant_id"):
                    tx["variant_title"] = v.get("title")
                    break
    
    return {"total": total, "transactions": transactions}

# ==================== ROLLBACK ROUTES (Admin Only) ====================

@api_router.post("/transactions/{transaction_id}/rollback")
async def rollback_transaction(transaction_id: str, current_user: dict = Depends(require_admin)):
    """Rollback a specific inventory transaction (admin only)"""
    tx = await db.inventory_transactions.find_one({"id": transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("is_rolled_back"):
        raise HTTPException(status_code=400, detail="Transaction already rolled back")
    
    tx_type = tx.get("transaction_type")
    variant_id = tx.get("variant_id")
    quantity = tx.get("quantity")
    
    # Reverse the transaction based on type
    if tx_type == TransactionType.RECEIVE:
        # Undo receive: subtract quantity from location/shelf
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("to_location_id"), "shelf_id": tx.get("to_shelf_id")},
            {"$inc": {"quantity": -quantity}}
        )
    elif tx_type == TransactionType.SALE:
        # Undo sale: add quantity back
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("from_location_id"), "shelf_id": tx.get("from_shelf_id")},
            {"$inc": {"quantity": quantity}}
        )
    elif tx_type == TransactionType.MOVE:
        # Undo move: move back from to_shelf to from_shelf
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("from_location_id"), "shelf_id": tx.get("to_shelf_id")},
            {"$inc": {"quantity": -quantity}}
        )
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("from_location_id"), "shelf_id": tx.get("from_shelf_id")},
            {"$inc": {"quantity": quantity}}
        )
    elif tx_type == TransactionType.TRANSFER:
        # Undo transfer: move back from to_location to from_location
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("to_location_id"), "shelf_id": tx.get("to_shelf_id")},
            {"$inc": {"quantity": -quantity}}
        )
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("from_location_id"), "shelf_id": tx.get("from_shelf_id")},
            {"$inc": {"quantity": quantity}}
        )
    elif tx_type == TransactionType.ADJUST:
        # Undo adjust: reverse the adjustment
        await db.inventory_levels.update_one(
            {"variant_id": variant_id, "location_id": tx.get("to_location_id"), "shelf_id": tx.get("to_shelf_id")},
            {"$inc": {"quantity": -quantity}}
        )
    
    # Mark transaction as rolled back
    await db.inventory_transactions.update_one(
        {"id": transaction_id},
        {"$set": {"is_rolled_back": True, "rolled_back_at": datetime.utcnow(), "rolled_back_by": current_user["id"]}}
    )
    
    # Get product info for log
    product = await db.products.find_one({"id": tx.get("product_id")})
    
    await log_action(
        ActionType.ROLLBACK,
        current_user,
        f"Rolled back {tx_type} transaction: {quantity}x {product.get('title', 'Unknown')}",
        entity_type="inventory_transaction",
        entity_id=transaction_id,
        old_data={"transaction": tx},
        metadata={"original_user_id": tx.get("user_id"), "original_username": tx.get("username")}
    )
    
    return {"message": "Transaction rolled back successfully"}

# ==================== LOCATION ROUTES ====================

@api_router.get("/locations", response_model=List[Location])
async def get_locations(current_user: dict = Depends(get_current_user)):
    locations = await db.locations.find({"is_active": True}, {"_id": 0}).to_list(100)
    return [Location(**loc) for loc in locations]

@api_router.post("/locations", response_model=Location)
async def create_location(data: LocationCreate, current_user: dict = Depends(require_admin)):
    location = Location(**data.dict())
    await db.locations.insert_one(location.dict())
    return location

@api_router.put("/locations/{location_id}", response_model=Location)
async def update_location(location_id: str, data: LocationCreate, current_user: dict = Depends(require_admin)):
    result = await db.locations.find_one_and_update(
        {"id": location_id},
        {"$set": data.dict()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Location not found")
    return Location(**result)

@api_router.delete("/locations/{location_id}")
async def delete_location(location_id: str, current_user: dict = Depends(require_admin)):
    await db.locations.update_one({"id": location_id}, {"$set": {"is_active": False}})
    return {"message": "Location deleted"}

# ==================== SHELF ROUTES ====================

@api_router.get("/shelves", response_model=List[Shelf])
async def get_shelves(location_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    if location_id:
        query["location_id"] = location_id
    shelves = await db.shelves.find(query, {"_id": 0}).to_list(1000)
    return [Shelf(**s) for s in shelves]

@api_router.get("/shelves/barcode/{barcode}", response_model=Shelf)
async def get_shelf_by_barcode(barcode: str, current_user: dict = Depends(get_current_user)):
    shelf = await db.shelves.find_one({"barcode": barcode, "is_active": True}, {"_id": 0})
    if not shelf:
        raise HTTPException(status_code=404, detail="Shelf not found")
    return Shelf(**shelf)

@api_router.post("/shelves", response_model=Shelf)
async def create_shelf(data: ShelfCreate, current_user: dict = Depends(require_admin)):
    # Check barcode uniqueness
    existing = await db.shelves.find_one({"barcode": data.barcode})
    if existing:
        raise HTTPException(status_code=400, detail="Barcode already exists")
    
    shelf = Shelf(**data.dict())
    await db.shelves.insert_one(shelf.dict())
    return shelf

@api_router.delete("/shelves/{shelf_id}")
async def delete_shelf(shelf_id: str, current_user: dict = Depends(require_admin)):
    await db.shelves.update_one({"id": shelf_id}, {"$set": {"is_active": False}})
    return {"message": "Shelf deleted"}

@api_router.put("/shelves/{shelf_id}", response_model=Shelf)
async def update_shelf(shelf_id: str, data: ShelfUpdate, current_user: dict = Depends(require_admin)):
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    # Check barcode uniqueness if changing
    if "barcode" in update_data:
        existing = await db.shelves.find_one({"barcode": update_data["barcode"], "id": {"$ne": shelf_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Barcode already exists")
    
    result = await db.shelves.find_one_and_update(
        {"id": shelf_id},
        {"$set": update_data},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Shelf not found")
    return Shelf(**result)

@api_router.get("/shelves/{shelf_id}", response_model=Shelf)
async def get_shelf(shelf_id: str, current_user: dict = Depends(get_current_user)):
    shelf = await db.shelves.find_one({"id": shelf_id, "is_active": True}, {"_id": 0})
    if not shelf:
        raise HTTPException(status_code=404, detail="Shelf not found")
    return Shelf(**shelf)

# ==================== PRODUCT ROUTES ====================

@api_router.get("/products")
async def get_products(
    search: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_active": True}
    if search:
        tokens = [t for t in re.split(r"\s+", search.strip()) if t]
        if tokens:
            token_clauses = []
            for t in tokens:
                rx = {"$regex": re.escape(t), "$options": "i"}
                token_clauses.append({
                    "$or": [
                        {"title": rx},
                        {"variants.title": rx},
                        {"variants.barcode": rx},
                        {"variants.upc_backup": rx},
                        {"variants.sku": rx}
                    ]
                })
            if len(tokens) > 1:
                combined = ".*".join([re.escape(t) for t in tokens])
                combo_rx = {"$regex": combined, "$options": "i"}
                query["$or"] = [
                    {"title": combo_rx},
                    {"variants.title": combo_rx},
                    {"variants.barcode": combo_rx},
                    {"variants.upc_backup": combo_rx},
                    {"variants.sku": combo_rx},
                    {"$and": token_clauses}
                ]
            else:
                query["$and"] = token_clauses
    
    total = await db.products.count_documents(query)
    # Keep list payload light on mobile: image_base64 can be very large.
    products = await db.products.find(query, {"_id": 0, "image_base64": 0}).skip(offset).limit(limit).to_list(limit)
    
    return {
        "products": products,
        "total": total,
        "limit": limit,
        "offset": offset
    }

@api_router.get("/products/barcode/{barcode}")
async def get_product_by_barcode(barcode: str, current_user: dict = Depends(get_current_user)):
    candidates = _barcode_candidates(barcode)
    if not candidates:
        candidates = [barcode]
    numeric_candidates = [int(c) for c in candidates if c.isdigit()]
    product = await db.products.find_one({
        "$or": [
            {"variants.barcode": {"$in": candidates}},
            {"variants.upc_backup": {"$in": candidates}},
            {"variants.sku": {"$in": candidates}},
            {"variants.barcode": {"$in": numeric_candidates}},
            {"variants.upc_backup": {"$in": numeric_candidates}},
            {"variants.barcode": barcode},
            {"variants.upc_backup": barcode},
            {"variants.sku": barcode},
        ],
        "is_active": True
    }, {"_id": 0})

    def _variant_matches(v: dict) -> bool:
        fields = [v.get("barcode"), v.get("upc_backup"), v.get("sku")]
        for field in fields:
            if not field:
                continue
            raw = str(field).strip()
            if raw == barcode or raw in candidates:
                return True
            norm = _normalize_barcode(raw)
            if norm == _normalize_barcode(barcode) or norm in candidates:
                return True
        return False

    if not product:
        # Fallback path: text search and normalized match in-memory.
        # This avoids false 404 when legacy data stores barcode formats differently.
        fallback_query = {
            "is_active": True,
            "$or": [
                {"variants.barcode": {"$regex": re.escape(barcode), "$options": "i"}},
                {"variants.upc_backup": {"$regex": re.escape(barcode), "$options": "i"}},
                {"variants.sku": {"$regex": re.escape(barcode), "$options": "i"}},
            ],
        }
        fallback_products = await db.products.find(fallback_query, {"_id": 0}).limit(200).to_list(200)
        product = next((p for p in fallback_products if any(_variant_matches(v) for v in p.get("variants", []))), None)
    if not product:
        # Last fallback: full normalized scan without hard limit.
        # Needed for stores with thousands of products and legacy numeric barcodes.
        cursor = db.products.find(
            {"is_active": True},
            {"_id": 0, "id": 1, "title": 1, "image_url": 1, "variants": 1}
        )
        async for p in cursor:
            if any(_variant_matches(v) for v in p.get("variants", [])):
                product = p
                break
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Find the specific variant
    variant = None
    for v in product.get("variants", []):
        if _variant_matches(v):
            variant = v
            break
    
    return {
        "product": product,
        "variant": variant
    }

@api_router.post("/stockx/lookup")
async def stockx_lookup(data: StockxLookupRequest, current_user: dict = Depends(get_current_user)):
    barcode = (data.barcode or "").strip()
    if not barcode:
        raise HTTPException(status_code=400, detail="Barcode required")

    lookup_id = str(uuid.uuid4())
    created_at = datetime.utcnow()
    normalized = _normalize_barcode(barcode)

    await db.stockx_lookups.insert_one({
        "id": lookup_id,
        "barcode": barcode,
        "normalized_barcode": normalized,
        "status": "pending",
        "match": None,
        "error": None,
        "requested_by": current_user.get("id"),
        "created_at": created_at,
        "updated_at": created_at,
    })

    try:
        override = await db.stockx_barcode_overrides.find_one({"normalized_barcode": normalized}, {"_id": 0})
        if not override:
            for cand in _barcode_candidates(barcode):
                override = await db.stockx_barcode_overrides.find_one({"normalized_barcode": cand}, {"_id": 0})
                if override:
                    break
        if override and override.get("match"):
            status_value = "found"
            await db.stockx_lookups.update_one(
                {"id": lookup_id},
                {"$set": {"status": status_value, "match": override.get("match"), "updated_at": datetime.utcnow()}}
            )
            return {"lookup_id": lookup_id, "status": status_value, "match": override.get("match")}
        match = await kicks_lookup_by_barcode(barcode)
        if isinstance(match, dict) and match.get("error"):
            status_value = "not_found"
            await db.stockx_lookups.update_one(
                {"id": lookup_id},
                {"$set": {"status": status_value, "match": None, "error": match.get("error"), "updated_at": datetime.utcnow()}}
            )
            return {"lookup_id": lookup_id, "status": status_value, "match": None, "error": match.get("error")}
        status_value = "found" if match else "not_found"
        if match:
            await _upsert_stockx_overrides_from_match(barcode, match)
        await db.stockx_lookups.update_one(
            {"id": lookup_id},
            {"$set": {"status": status_value, "match": match, "updated_at": datetime.utcnow()}}
        )
        return {"lookup_id": lookup_id, "status": status_value, "match": match}
    except Exception as exc:
        await db.stockx_lookups.update_one(
            {"id": lookup_id},
            {"$set": {"status": "error", "error": str(exc), "updated_at": datetime.utcnow()}}
        )
        return {"lookup_id": lookup_id, "status": "error", "match": None, "error": str(exc)}

@api_router.post("/stockx/lookup-url")
async def stockx_lookup_url(data: StockxLookupUrlRequest, current_user: dict = Depends(get_current_user)):
    url = (data.url or "").strip()
    slug = _stockx_slug_from_url(url)
    if not slug:
        raise HTTPException(status_code=400, detail="Invalid StockX URL")

    full = await _kicks_get_product_full(slug)
    if not full:
        raise HTTPException(status_code=404, detail="Product not found on StockX")

    stockx_url = _stockx_url_from_product(full) or url
    match = {
        "title": full.get("title") or slug,
        "brand": full.get("brand") or full.get("vendor"),
        "slug": slug,
        "image_url": full.get("image") or full.get("imageUrl") or (full.get("images") or [None])[0],
        "stockx_url": stockx_url,
        "verified": True,
    }
    barcode = (data.barcode or "").strip()
    if barcode:
        await _upsert_stockx_overrides_from_match(barcode, match)
    return match

@api_router.post("/stockx/lookup-query")
async def stockx_lookup_query(data: StockxLookupQueryRequest, current_user: dict = Depends(get_current_user)):
    query = (data.query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query required")
    match = await kicks_lookup_by_query(query)
    if not match:
        raise HTTPException(status_code=404, detail="Product not found on StockX")
    return match

@api_router.get("/stockx/lookup")
async def list_stockx_lookups(status: Optional[str] = None, skip: int = 0, limit: int = 50, current_user: dict = Depends(require_admin)):
    query: dict = {}
    if status:
        query["status"] = status
    items = await db.stockx_lookups.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.stockx_lookups.count_documents(query)
    return {"items": items, "total": total, "limit": limit, "offset": skip}

def _build_stockx_import_env(data: StockxImportRequest) -> dict:
    env: dict = {}
    env["STOCKX_PRODUCT_URL"] = data.stockx_url
    env["STOCKX_URL"] = data.stockx_url
    env["PRODUCT_TYPE"] = data.product_type
    if data.tags:
        env["TAGS"] = data.tags
    if data.template_suffix:
        env["TEMPLATE_SUFFIX"] = data.template_suffix
    if data.status:
        env["STATUS"] = data.status
    if data.default_qty is not None:
        env["DEFAULT_QTY"] = str(int(data.default_qty))

    if data.express_mode == "all":
        env["SIZE_SUFFIX_MODE"] = "express"
        env["EXPRESS_LABEL"] = data.express_label or "Express 24/48H"
    else:
        env["SIZE_SUFFIX_MODE"] = "off"

    if data.price_mode == "fixed":
        if data.fixed_price_eur is None:
            raise HTTPException(status_code=400, detail="Fixed price required for fixed price mode")
        env["FIXED_PRICE_EUR"] = str(int(float(data.fixed_price_eur)))
    else:
        env["PRICE_MODE"] = "market"
        env["PRICE_TYPE_PREFERRED"] = (data.price_type_preferred or "standard").lower()
        env["ALLOW_UNPRICED"] = "0"
        if data.fixed_price_eur is not None:
            env["FIXED_PRICE_EUR"] = ""
    if data.price_mode == "fixed":
        env["PRICE_MODE"] = "fixed"

    if data.size_mode == "range":
        if data.eu_min is not None:
            env["EU_MIN"] = str(data.eu_min)
        if data.eu_max is not None:
            env["EU_MAX"] = str(data.eu_max)
    elif data.size_mode == "list":
        if data.eu_list:
            env["EU_LIST"] = ",".join([str(x) for x in data.eu_list])
    return env

async def _delayed_sync_new(current_user: dict, job_id: Optional[str] = None):
    await asyncio.sleep(120)
    if not shopify_sync_lock.locked():
        await run_shopify_sync("new", current_user)

    if job_id:
        job = await db.stockx_import_jobs.find_one({"id": job_id}, {"_id": 0})
        if not job:
            return
        options = job.get("options") or {}
        if (options.get("price_mode") or "fixed") == "market":
            shopify_product_id = job.get("shopify_product_id")
            stockx_url = job.get("stockx_url") or options.get("stockx_url")
            if shopify_product_id and stockx_url:
                express_mode = options.get("express_mode") or "none"
                express_label = options.get("express_label") or "Express 24/48H"
                added = await _add_unpriced_variants_to_wms(
                    shopify_product_id=str(shopify_product_id),
                    stockx_url=stockx_url,
                    express_mode=express_mode,
                    express_label=express_label
                )
                await db.stockx_import_jobs.update_one(
                    {"id": job_id},
                    {"$set": {"unpriced_variants_added": added, "updated_at": datetime.utcnow()}}
                )

async def _run_stockx_import_job(job_id: str, env_overrides: dict, current_user: dict):
    await db.stockx_import_jobs.update_one(
        {"id": job_id},
        {"$set": {"status": "running", "started_at": datetime.utcnow(), "updated_at": datetime.utcnow()}}
    )

    env = os.environ.copy()
    for k, v in env_overrides.items():
        if v is not None:
            env[k] = str(v)

    cmd = [sys.executable, str(ROOT_DIR / "IMPORTAPRODOTTISTOCKX.PY")]
    try:
        result = await asyncio.to_thread(
            subprocess.run,
            cmd,
            env=env,
            capture_output=True,
            text=True,
            timeout=3600
        )
        status_value = "success" if result.returncode == 0 else "error"
        shopify_product_id = None
        m = re.search(r"Created product_id=\s*(\d+)", result.stdout or "")
        if m:
            shopify_product_id = m.group(1)
        await db.stockx_import_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": status_value,
                "stdout": result.stdout[-20000:],
                "stderr": result.stderr[-20000:],
                "return_code": result.returncode,
                "shopify_product_id": shopify_product_id,
                "completed_at": datetime.utcnow(),
                "updated_at": datetime.utcnow(),
            }}
        )
        if status_value == "success":
            asyncio.create_task(_delayed_sync_new(current_user, job_id))
    except Exception as exc:
        await db.stockx_import_jobs.update_one(
            {"id": job_id},
            {"$set": {"status": "error", "error": str(exc), "completed_at": datetime.utcnow(), "updated_at": datetime.utcnow()}}
        )

@api_router.post("/stockx/import")
async def stockx_import(data: StockxImportRequest, background_tasks: BackgroundTasks, current_user: dict = Depends(require_admin)):
    if not data.stockx_url:
        raise HTTPException(status_code=400, detail="stockx_url required")

    env_overrides = _build_stockx_import_env(data)

    job_id = str(uuid.uuid4())
    created_at = datetime.utcnow()

    await db.stockx_import_jobs.insert_one({
        "id": job_id,
        "barcode": data.barcode,
        "stockx_url": data.stockx_url,
        "options": data.dict(),
        "status": "queued",
        "requested_by": current_user.get("id"),
        "created_at": created_at,
        "updated_at": created_at,
    })

    if data.lookup_id:
        await db.stockx_lookups.update_one(
            {"id": data.lookup_id},
            {"$set": {"import_job_id": job_id, "updated_at": datetime.utcnow()}}
        )

    background_tasks.add_task(_run_stockx_import_job, job_id, env_overrides, current_user)
    return {"job_id": job_id, "status": "queued"}

@api_router.get("/stockx/imports")
async def list_stockx_imports(status: Optional[str] = None, skip: int = 0, limit: int = 50, current_user: dict = Depends(require_admin)):
    query: dict = {}
    if status:
        query["status"] = status
    items = await db.stockx_import_jobs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.stockx_import_jobs.count_documents(query)
    return {"items": items, "total": total, "limit": limit, "offset": skip}

@api_router.get("/products/{product_id}")
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id, "is_active": True}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

@api_router.post("/products")
async def create_product(data: ProductCreate, current_user: dict = Depends(require_admin)):
    product_id = str(uuid.uuid4())
    variants = []
    for v in data.variants:
        variants.append(ProductVariant(
            product_id=product_id,
            title=v.get("title", "Default"),
            sku=v.get("sku"),
            barcode=v.get("barcode"),
            price=v.get("price")
        ).dict())
    
    product = Product(
        id=product_id,
        title=data.title,
        image_url=data.image_url,
        variants=variants
    )
    await db.products.insert_one(product.dict())
    return product.dict()

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, data: ProductUpdate, current_user: dict = Depends(require_admin)):
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.products.find_one_and_update(
        {"id": product_id},
        {"$set": update_data},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Product not found")
    return result

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(require_admin)):
    await db.products.update_one({"id": product_id}, {"$set": {"is_active": False}})
    return {"message": "Product deleted"}

@api_router.put("/products/{product_id}/variants/{variant_id}")
async def update_variant(product_id: str, variant_id: str, data: VariantUpdate, current_user: dict = Depends(require_admin)):
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Update the specific variant
    variants = product.get("variants", [])
    variant_found = False
    for i, v in enumerate(variants):
        if v.get("id") == variant_id:
            update_data = {k: val for k, val in data.dict().items() if val is not None}
            variants[i] = {**v, **update_data}
            variant_found = True
            break
    
    if not variant_found:
        raise HTTPException(status_code=404, detail="Variant not found")
    
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"variants": variants, "updated_at": datetime.utcnow()}}
    )
    
    return {"message": "Variant updated", "variant": variants[i]}

# ==================== COLLECTION ROUTES ====================

@api_router.get("/collections")
async def get_collections(current_user: dict = Depends(get_current_user)):
    collections = await db.collections.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return collections

@api_router.get("/collections/{collection_id}")
async def get_collection(collection_id: str, current_user: dict = Depends(get_current_user)):
    collection = await db.collections.find_one({"id": collection_id, "is_active": True}, {"_id": 0})
    if not collection:
        raise HTTPException(status_code=404, detail="Collection not found")
    return collection

@api_router.get("/collections/{collection_id}/products")
async def get_collection_products(collection_id: str, current_user: dict = Depends(get_current_user)):
    collection = await db.collections.find_one({"id": collection_id, "is_active": True}, {"_id": 0})
    if not collection:
        raise HTTPException(status_code=404, detail="Collection not found")
    
    product_ids = collection.get("product_ids", [])
    products = await db.products.find({"id": {"$in": product_ids}, "is_active": True}, {"_id": 0}).to_list(1000)
    return products

@api_router.post("/collections")
async def create_collection(data: CollectionCreate, current_user: dict = Depends(require_admin)):
    collection = Collection(**data.dict())
    await db.collections.insert_one(collection.dict())
    return collection.dict()

@api_router.put("/collections/{collection_id}")
async def update_collection(collection_id: str, data: CollectionCreate, current_user: dict = Depends(require_admin)):
    result = await db.collections.find_one_and_update(
        {"id": collection_id},
        {"$set": data.dict()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Collection not found")
    return result

@api_router.delete("/collections/{collection_id}")
async def delete_collection(collection_id: str, current_user: dict = Depends(require_admin)):
    await db.collections.update_one({"id": collection_id}, {"$set": {"is_active": False}})
    return {"message": "Collection deleted"}

# ==================== INVENTORY ROUTES ====================

async def _get_fallback_location(location_id: Optional[str]):
    if location_id:
        return await db.locations.find_one({"id": location_id}, {"_id": 0})

    location = await db.locations.find_one(
        {"name": {"$regex": "warehouse|magazzino", "$options": "i"}, "is_active": True},
        {"_id": 0}
    )
    if not location:
        location = await db.locations.find_one({"is_active": True}, {"_id": 0})
    return location


async def _get_inventory_levels_with_fallback(location_id: Optional[str]):
    query = {}
    if location_id:
        query["location_id"] = location_id

    inventory_raw = await db.inventory_levels.find(query, {"_id": 0}).to_list(10000)
    inventory = []
    for inv in inventory_raw:
        try:
            qty = int(inv.get("quantity", 0))
        except (TypeError, ValueError):
            qty = 0
        if qty > 0:
            inv["quantity"] = qty
            inventory.append(inv)
    variant_ids = {inv.get("variant_id") for inv in inventory if inv.get("variant_id")}

    fallback_location = await _get_fallback_location(location_id)
    if not fallback_location:
        return inventory

    products = await db.products.find({"is_active": True}, {"_id": 0}).to_list(10000)
    for product in products:
        for variant in product.get("variants", []):
            variant_id = variant.get("id")
            if not variant_id or variant_id in variant_ids:
                continue

            qty_raw = variant.get("inventory_quantity")
            if qty_raw is None:
                continue
            try:
                qty = int(qty_raw)
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue

            inventory.append({
                "id": f"fallback-{variant_id}",
                "variant_id": variant_id,
                "product_id": product.get("id"),
                "location_id": fallback_location.get("id"),
                "shelf_id": None,
                "quantity": qty,
                "updated_at": datetime.utcnow(),
                "source": "shopify_fallback",
                "last_synced_at": None
            })
            variant_ids.add(variant_id)

    return inventory


@api_router.get("/inventory")
async def get_inventory(
    location_id: Optional[str] = None,
    shelf_id: Optional[str] = None,
    variant_id: Optional[str] = None,
    limit: int = 0,
    offset: int = 0,
    paged: bool = False,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if location_id:
        query["location_id"] = location_id
    if shelf_id:
        query["shelf_id"] = shelf_id
    if variant_id:
        query["variant_id"] = variant_id

    if paged or limit or offset:
        total = await db.inventory_levels.count_documents(query)
        limit = limit if limit > 0 else 50
        items = await db.inventory_levels.find(query, {"_id": 0}).skip(offset).limit(limit).to_list(limit)
        return {"items": items, "total": total, "limit": limit, "offset": offset}

    inventory = await db.inventory_levels.find(query, {"_id": 0}).to_list(10000)
    return inventory

@api_router.get("/inventory/summary")
async def get_inventory_summary(
    location_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get inventory summary with product details"""
    inventory = await _get_inventory_levels_with_fallback(location_id)

    if not inventory:
        return []

    variant_ids = {inv.get("variant_id") for inv in inventory if inv.get("variant_id")}
    product_ids = {inv.get("product_id") for inv in inventory if inv.get("product_id")}
    location_ids = {inv.get("location_id") for inv in inventory if inv.get("location_id")}
    shelf_ids = {inv.get("shelf_id") for inv in inventory if inv.get("shelf_id")}

    products = await db.products.find(
        {
            "is_active": True,
            "$or": [
                {"id": {"$in": list(product_ids)}},
                {"variants.id": {"$in": list(variant_ids)}},
            ],
        },
        {
            "_id": 0,
            "id": 1,
            "title": 1,
            "image_base64": 1,
            "image_url": 1,
            "variants.id": 1,
            "variants.title": 1,
            "variants.barcode": 1,
            "variants.sku": 1,
        },
    ).to_list(10000)

    product_by_id = {p.get("id"): p for p in products if p.get("id")}
    variant_map = {}
    for p in products:
        for v in p.get("variants", []):
            vid = v.get("id")
            if vid:
                variant_map[vid] = (p, v)

    locations = await db.locations.find({"id": {"$in": list(location_ids)}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    shelves = await db.shelves.find({"id": {"$in": list(shelf_ids)}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    location_by_id = {l.get("id"): l for l in locations if l.get("id")}
    shelf_by_id = {s.get("id"): s for s in shelves if s.get("id")}

    result = []
    for inv in inventory:
        product = None
        variant = None
        if inv.get("variant_id") in variant_map:
            product, variant = variant_map.get(inv.get("variant_id"))
        if not product and inv.get("product_id"):
            product = product_by_id.get(inv.get("product_id"))
            if product:
                variant = next((v for v in product.get("variants", []) if v.get("id") == inv.get("variant_id")), None)
        location = location_by_id.get(inv.get("location_id"))
        shelf = shelf_by_id.get(inv.get("shelf_id"))
        fallback_title = f"Unknown ({inv.get('variant_id')})"

        result.append({
            **inv,
            "product_title": product.get("title") if product else fallback_title,
            "product_image": (product.get("image_base64") or product.get("image_url")) if product else None,
            "variant_title": variant.get("title") if variant else None,
            "variant_barcode": variant.get("barcode") if variant else None,
            "variant_upc_backup": variant.get("upc_backup") if variant else None,
            "variant_sku": variant.get("sku") if variant else None,
            "location_name": location.get("name") if location else None,
            "shelf_name": shelf.get("name") if shelf else None
        })

    return result

# ==================== INVENTORY OPERATIONS ====================

@api_router.post("/inventory/receive")
async def receive_inventory(data: ReceiveRequest, current_user: dict = Depends(get_current_user)):
    """Receive inventory (add stock)"""
    # Find variant and product
    product = await db.products.find_one({"variants.id": data.variant_id})
    if not product:
        raise HTTPException(status_code=404, detail="Variant not found")
    
    # Get variant info
    variant_title = ""
    for v in product.get("variants", []):
        if v.get("id") == data.variant_id:
            variant_title = v.get("title", "")
            break
    
    # Update or create inventory level
    existing = await db.inventory_levels.find_one({
        "variant_id": data.variant_id,
        "location_id": data.location_id,
        "shelf_id": data.shelf_id
    })
    
    if existing:
        new_qty = existing["quantity"] + data.quantity
        await db.inventory_levels.update_one(
            {"id": existing["id"]},
            {"$set": {"quantity": new_qty, "updated_at": datetime.utcnow(), "product_id": product["id"]}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=data.variant_id,
            product_id=product["id"],
            location_id=data.location_id,
            shelf_id=data.shelf_id,
            quantity=data.quantity,
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())
    
    # Create transaction with username
    tx = InventoryTransaction(
        transaction_type=TransactionType.RECEIVE,
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=data.quantity,
        to_location_id=data.location_id,
        to_shelf_id=data.shelf_id,
        user_id=current_user["id"],
        username=current_user["username"]
    )
    await db.inventory_transactions.insert_one(tx.dict())

    # Keep Shopify quantity aligned with local WMS quantity changes.
    shopify_updated = await update_shopify_inventory(data.variant_id, data.quantity, current_user)
    
    # Log action
    await log_action(
        ActionType.INVENTORY_RECEIVE,
        current_user,
        f"Received {data.quantity}x {product.get('title')} ({variant_title})" + (" [Shopify updated]" if shopify_updated else ""),
        entity_type="inventory",
        entity_id=tx.id,
        new_data={"quantity": data.quantity, "location_id": data.location_id, "shelf_id": data.shelf_id, "shopify_synced": shopify_updated}
    )
    
    return {"message": "Inventory received", "transaction_id": tx.id, "shopify_updated": shopify_updated}

@api_router.post("/inventory/move")
async def move_inventory(data: MoveRequest, current_user: dict = Depends(get_current_user)):
    """Move inventory between shelves in the same location"""
    # Check source inventory
    source = await db.inventory_levels.find_one({
        "variant_id": data.variant_id,
        "location_id": data.location_id,
        "shelf_id": data.from_shelf_id
    })
    
    if not source or source["quantity"] < data.quantity:
        raise HTTPException(status_code=400, detail="Insufficient inventory")
    
    product = await db.products.find_one({"variants.id": data.variant_id})
    
    # Decrease source
    new_source_qty = source["quantity"] - data.quantity
    await db.inventory_levels.update_one(
        {"id": source["id"]},
        {"$set": {"quantity": new_source_qty, "updated_at": datetime.utcnow()}}
    )
    
    # Increase destination
    dest = await db.inventory_levels.find_one({
        "variant_id": data.variant_id,
        "location_id": data.location_id,
        "shelf_id": data.to_shelf_id
    })
    
    if dest:
        new_dest_qty = dest["quantity"] + data.quantity
        await db.inventory_levels.update_one(
            {"id": dest["id"]},
            {"$set": {"quantity": new_dest_qty, "updated_at": datetime.utcnow(), "product_id": product["id"]}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=data.variant_id,
            product_id=product["id"],
            location_id=data.location_id,
            shelf_id=data.to_shelf_id,
            quantity=data.quantity,
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())
    
    # Create transaction
    tx = InventoryTransaction(
        transaction_type=TransactionType.MOVE,
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=data.quantity,
        from_location_id=data.location_id,
        to_location_id=data.location_id,
        from_shelf_id=data.from_shelf_id,
        to_shelf_id=data.to_shelf_id,
        user_id=current_user["id"]
    )
    await db.inventory_transactions.insert_one(tx.dict())
    
    return {"message": "Inventory moved", "transaction_id": tx.id}

@api_router.post("/inventory/transfer")
async def transfer_inventory(data: TransferRequest, current_user: dict = Depends(get_current_user)):
    """Transfer inventory between locations"""
    # Check source inventory
    source_query = {
        "variant_id": data.variant_id,
        "location_id": data.from_location_id
    }
    if data.from_shelf_id:
        source_query["shelf_id"] = data.from_shelf_id
    
    source = await db.inventory_levels.find_one(source_query)
    
    if not source or source["quantity"] < data.quantity:
        raise HTTPException(status_code=400, detail="Insufficient inventory")
    
    product = await db.products.find_one({"variants.id": data.variant_id})
    
    # Decrease source
    new_source_qty = source["quantity"] - data.quantity
    await db.inventory_levels.update_one(
        {"id": source["id"]},
        {"$set": {"quantity": new_source_qty, "updated_at": datetime.utcnow()}}
    )
    
    # Increase destination
    dest_query = {
        "variant_id": data.variant_id,
        "location_id": data.to_location_id
    }
    if data.to_shelf_id:
        dest_query["shelf_id"] = data.to_shelf_id
    
    dest = await db.inventory_levels.find_one(dest_query)
    
    if dest:
        new_dest_qty = dest["quantity"] + data.quantity
        await db.inventory_levels.update_one(
            {"id": dest["id"]},
            {"$set": {"quantity": new_dest_qty, "updated_at": datetime.utcnow(), "product_id": product["id"]}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=data.variant_id,
            product_id=product["id"],
            location_id=data.to_location_id,
            shelf_id=data.to_shelf_id,
            quantity=data.quantity,
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())
    
    # Create transaction
    tx = InventoryTransaction(
        transaction_type=TransactionType.TRANSFER,
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=data.quantity,
        from_location_id=data.from_location_id,
        to_location_id=data.to_location_id,
        from_shelf_id=data.from_shelf_id,
        to_shelf_id=data.to_shelf_id,
        user_id=current_user["id"]
    )
    await db.inventory_transactions.insert_one(tx.dict())
    
    return {"message": "Inventory transferred", "transaction_id": tx.id}

@api_router.post("/inventory/sale")
async def sale_inventory(data: SaleRequest, current_user: dict = Depends(get_current_user)):
    """Record a sale (decrease inventory) - Also updates Shopify"""
    # Check source inventory
    source_query = {
        "variant_id": data.variant_id,
        "location_id": data.location_id
    }
    if data.shelf_id:
        source_query["shelf_id"] = data.shelf_id
    
    source = await db.inventory_levels.find_one(source_query)
    
    if not source or source["quantity"] < data.quantity:
        raise HTTPException(status_code=400, detail="Insufficient inventory")
    
    product = await db.products.find_one({"variants.id": data.variant_id})
    
    # Get variant title
    variant_title = ""
    for v in product.get("variants", []):
        if v.get("id") == data.variant_id:
            variant_title = v.get("title", "")
            break
    
    # Decrease inventory
    new_qty = source["quantity"] - data.quantity
    await db.inventory_levels.update_one(
        {"id": source["id"]},
        {"$set": {"quantity": new_qty, "updated_at": datetime.utcnow()}}
    )
    
    # Create transaction with username
    tx = InventoryTransaction(
        transaction_type=TransactionType.SALE,
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=data.quantity,
        from_location_id=data.location_id,
        from_shelf_id=data.shelf_id,
        sale_price=data.sale_price,
        user_id=current_user["id"],
        username=current_user["username"]
    )
    await db.inventory_transactions.insert_one(tx.dict())
    
    # Update Shopify inventory (subtract quantity)
    shopify_updated = await update_shopify_inventory(data.variant_id, -data.quantity, current_user)
    
    # Log action
    await log_action(
        ActionType.INVENTORY_SALE,
        current_user,
        f"Sold {data.quantity}x {product.get('title')} ({variant_title}) at EUR {data.sale_price}" + (" [Shopify updated]" if shopify_updated else ""),
        entity_type="inventory",
        entity_id=tx.id,
        new_data={"quantity": data.quantity, "sale_price": data.sale_price, "shopify_synced": shopify_updated}
    )
    
    return {"message": "Sale recorded", "transaction_id": tx.id, "shopify_updated": shopify_updated}

@api_router.post("/inventory/adjust")
async def adjust_inventory(data: AdjustRequest, current_user: dict = Depends(require_admin)):
    """Adjust inventory (admin only)"""
    product = await db.products.find_one({"variants.id": data.variant_id})
    if not product:
        raise HTTPException(status_code=404, detail="Variant not found")
    
    # Find or create inventory level
    query = {
        "variant_id": data.variant_id,
        "location_id": data.location_id
    }
    if data.shelf_id:
        query["shelf_id"] = data.shelf_id
    
    existing = await db.inventory_levels.find_one(query)
    old_quantity = existing["quantity"] if existing else 0
    quantity_diff = data.new_quantity - old_quantity
    
    if existing:
        await db.inventory_levels.update_one(
            {"id": existing["id"]},
            {"$set": {"quantity": data.new_quantity, "updated_at": datetime.utcnow(), "product_id": product["id"]}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=data.variant_id,
            product_id=product["id"],
            location_id=data.location_id,
            shelf_id=data.shelf_id,
            quantity=data.new_quantity,
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())
    
    # Create transaction
    tx = InventoryTransaction(
        transaction_type=TransactionType.ADJUST,
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=quantity_diff,
        to_location_id=data.location_id,
        to_shelf_id=data.shelf_id,
        note=data.note or f"Adjusted from {old_quantity} to {data.new_quantity}",
        user_id=current_user["id"]
    )
    await db.inventory_transactions.insert_one(tx.dict())
    
    return {"message": "Inventory adjusted", "transaction_id": tx.id}

# ==================== STALLO (Temporary Out) ====================

@api_router.get("/stall-items")
async def get_stall_items(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    items = await db.stall_items.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    if not items:
        return []

    variant_ids = {i.get("variant_id") for i in items if i.get("variant_id")}
    product_ids = {i.get("product_id") for i in items if i.get("product_id")}
    location_ids = {i.get("from_location_id") for i in items if i.get("from_location_id")}
    location_ids.update({i.get("to_location_id") for i in items if i.get("to_location_id")})
    shelf_ids = {i.get("from_shelf_id") for i in items if i.get("from_shelf_id")}
    shelf_ids.update({i.get("to_shelf_id") for i in items if i.get("to_shelf_id")})

    products = await db.products.find(
        {
            "$or": [
                {"id": {"$in": list(product_ids)}},
                {"variants.id": {"$in": list(variant_ids)}},
            ],
        },
        {
            "_id": 0,
            "id": 1,
            "title": 1,
            "variants.id": 1,
            "variants.title": 1,
            "variants.barcode": 1,
        },
    ).to_list(10000)

    product_by_id = {p.get("id"): p for p in products if p.get("id")}
    variant_map = {}
    for p in products:
        for v in p.get("variants", []):
            vid = v.get("id")
            if vid:
                variant_map[vid] = (p, v)

    locations = await db.locations.find({"id": {"$in": list(location_ids)}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    shelves = await db.shelves.find({"id": {"$in": list(shelf_ids)}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    location_by_id = {l.get("id"): l for l in locations if l.get("id")}
    shelf_by_id = {s.get("id"): s for s in shelves if s.get("id")}

    enriched = []
    for item in items:
        product = None
        variant = None
        if item.get("variant_id") in variant_map:
            product, variant = variant_map.get(item.get("variant_id"))
        if not product and item.get("product_id"):
            product = product_by_id.get(item.get("product_id"))
            if product:
                variant = next((v for v in product.get("variants", []) if v.get("id") == item.get("variant_id")), None)
        from_loc = location_by_id.get(item.get("from_location_id"))
        from_shelf = shelf_by_id.get(item.get("from_shelf_id"))
        to_loc = location_by_id.get(item.get("to_location_id"))
        to_shelf = shelf_by_id.get(item.get("to_shelf_id"))

        enriched.append({
            **item,
            "product_title": product.get("title") if product else None,
            "variant_title": variant.get("title") if variant else None,
            "variant_barcode": variant.get("barcode") if variant else None,
            "variant_upc_backup": variant.get("upc_backup") if variant else None,
            "from_location_name": from_loc.get("name") if from_loc else None,
            "from_shelf_name": from_shelf.get("name") if from_shelf else None,
            "to_location_name": to_loc.get("name") if to_loc else None,
            "to_shelf_name": to_shelf.get("name") if to_shelf else None,
        })

    return enriched

@api_router.post("/stall-items")
async def create_stall_item(data: StallCreateRequest, current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"variants.id": data.variant_id})
    if not product:
        raise HTTPException(status_code=404, detail="Variant not found")

    # Check inventory availability
    source_query = {
        "variant_id": data.variant_id,
        "location_id": data.location_id
    }
    if data.shelf_id:
        source_query["shelf_id"] = data.shelf_id
    source = await db.inventory_levels.find_one(source_query)
    if not source or source["quantity"] < data.quantity:
        raise HTTPException(status_code=400, detail="Insufficient inventory")

    # Decrease inventory at source
    new_source_qty = source["quantity"] - data.quantity
    await db.inventory_levels.update_one(
        {"id": source["id"]},
        {"$set": {"quantity": new_source_qty, "updated_at": datetime.utcnow(), "product_id": product["id"]}}
    )

    stall_item = StallItem(
        variant_id=data.variant_id,
        product_id=product["id"],
        quantity=data.quantity,
        from_location_id=data.location_id,
        from_shelf_id=data.shelf_id,
        note=data.note,
        customer_name=data.customer_name,
        created_by=current_user.get("username")
    )
    await db.stall_items.insert_one(stall_item.dict())

    await log_action(
        ActionType.STALL_CREATE,
        current_user,
        f"Stallo: {data.quantity}x {product.get('title')}",
        entity_type="stall",
        entity_id=stall_item.id,
        new_data=stall_item.dict()
    )

    return stall_item.dict()

@api_router.post("/stall-items/{stall_id}/return")
async def return_stall_item(stall_id: str, data: StallReturnRequest, current_user: dict = Depends(get_current_user)):
    item = await db.stall_items.find_one({"id": stall_id})
    if not item:
        raise HTTPException(status_code=404, detail="Stall item not found")
    if item.get("status") != "in_stallo":
        raise HTTPException(status_code=400, detail="Stall item not active")

    to_location_id = data.to_location_id or item.get("from_location_id")
    to_shelf_id = data.to_shelf_id if data.to_shelf_id is not None else item.get("from_shelf_id")

    # Increase inventory at destination
    dest_query = {"variant_id": item["variant_id"], "location_id": to_location_id}
    if to_shelf_id:
        dest_query["shelf_id"] = to_shelf_id
    dest = await db.inventory_levels.find_one(dest_query)
    if dest:
        await db.inventory_levels.update_one(
            {"id": dest["id"]},
            {"$set": {"quantity": dest["quantity"] + item["quantity"], "updated_at": datetime.utcnow(), "product_id": item.get("product_id")}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=item["variant_id"],
            product_id=item.get("product_id"),
            location_id=to_location_id,
            shelf_id=to_shelf_id,
            quantity=item["quantity"],
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())

    await db.stall_items.update_one(
        {"id": stall_id},
        {"$set": {
            "status": "returned",
            "resolved_action": "returned",
            "resolved_at": datetime.utcnow(),
            "to_location_id": to_location_id,
            "to_shelf_id": to_shelf_id
        }}
    )

    await log_action(
        ActionType.STALL_RETURN,
        current_user,
        f"Stallo rientrato: {item.get('quantity')}x",
        entity_type="stall",
        entity_id=stall_id,
        new_data={"to_location_id": to_location_id, "to_shelf_id": to_shelf_id}
    )

    return {"message": "Stall item returned"}

@api_router.post("/stall-items/{stall_id}/sell")
async def sell_stall_item(stall_id: str, data: StallSellRequest, current_user: dict = Depends(get_current_user)):
    item = await db.stall_items.find_one({"id": stall_id})
    if not item:
        raise HTTPException(status_code=404, detail="Stall item not found")
    if item.get("status") != "in_stallo":
        raise HTTPException(status_code=400, detail="Stall item not active")

    await db.stall_items.update_one(
        {"id": stall_id},
        {"$set": {
            "status": "sold",
            "resolved_action": "sold",
            "resolved_at": datetime.utcnow()
        }}
    )

    await log_action(
        ActionType.STALL_SELL,
        current_user,
        f"Stallo venduto: {item.get('quantity')}x",
        entity_type="stall",
        entity_id=stall_id,
        new_data={"sale_price": data.sale_price}
    )

    return {"message": "Stall item sold"}

@api_router.post("/stall-items/{stall_id}/move")
async def move_stall_item(stall_id: str, data: StallMoveRequest, current_user: dict = Depends(get_current_user)):
    item = await db.stall_items.find_one({"id": stall_id})
    if not item:
        raise HTTPException(status_code=404, detail="Stall item not found")
    if item.get("status") != "in_stallo":
        raise HTTPException(status_code=400, detail="Stall item not active")

    to_location_id = data.to_location_id
    to_shelf_id = data.to_shelf_id

    dest_query = {"variant_id": item["variant_id"], "location_id": to_location_id}
    if to_shelf_id:
        dest_query["shelf_id"] = to_shelf_id
    dest = await db.inventory_levels.find_one(dest_query)
    if dest:
        await db.inventory_levels.update_one(
            {"id": dest["id"]},
            {"$set": {"quantity": dest["quantity"] + item["quantity"], "updated_at": datetime.utcnow(), "product_id": item.get("product_id")}}
        )
    else:
        inv_level = InventoryLevel(
            variant_id=item["variant_id"],
            product_id=item.get("product_id"),
            location_id=to_location_id,
            shelf_id=to_shelf_id,
            quantity=item["quantity"],
            source="local"
        )
        await db.inventory_levels.insert_one(inv_level.dict())

    await db.stall_items.update_one(
        {"id": stall_id},
        {"$set": {
            "status": "moved",
            "resolved_action": "moved",
            "resolved_at": datetime.utcnow(),
            "to_location_id": to_location_id,
            "to_shelf_id": to_shelf_id
        }}
    )

    await log_action(
        ActionType.STALL_MOVE,
        current_user,
        f"Stallo spostato: {item.get('quantity')}x",
        entity_type="stall",
        entity_id=stall_id,
        new_data={"to_location_id": to_location_id, "to_shelf_id": to_shelf_id}
    )

    return {"message": "Stall item moved"}

# ==================== TRANSACTION HISTORY ====================

@api_router.get("/transactions")
async def get_transactions(
    limit: int = 50,
    offset: int = 0,
    transaction_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    total = await db.inventory_transactions.count_documents(query)
    transactions = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(offset).limit(limit).to_list(limit)
    
    # Enrich with product data
    enriched = []
    for tx in transactions:
        product = await db.products.find_one({"id": tx["product_id"]})
        variant = None
        if product:
            variant = next((v for v in product.get("variants", []) if v["id"] == tx["variant_id"]), None)
        
        enriched.append({
            **tx,
            "product_title": product.get("title") if product else None,
            "variant_title": variant.get("title") if variant else None,
            "variant_barcode": variant.get("barcode") if variant else None,
            "variant_upc_backup": variant.get("upc_backup") if variant else None
        })
    
    return {
        "transactions": enriched,
        "total": total,
        "limit": limit,
        "offset": offset
    }

# ==================== SHOPIFY INTEGRATION ====================

async def download_image_as_base64(url: str) -> Optional[str]:
    """Download image and convert to base64"""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as response:
                if response.status == 200:
                    image_data = await response.read()
                    base64_data = base64.b64encode(image_data).decode('utf-8')
                    content_type = response.headers.get('content-type', 'image/jpeg')
                    return f"data:{content_type};base64,{base64_data}"
    except Exception as e:
        logger.error(f"Error downloading image: {e}")
    return None

@api_router.post("/shopify/sync")
async def sync_shopify_products(current_user: dict = Depends(require_admin)):
    """AGGIORNA TUTTI I PRODOTTI - Avvia sync in background (READ ONLY)"""
    if shopify_sync_lock.locked():
        return {"status": "syncing", "message": "Sincronizzazione giÃ  in corso"}
    asyncio.create_task(run_shopify_sync("all", current_user))
    return {"status": "syncing", "message": "Sincronizzazione avviata"}

async def run_shopify_sync(mode: str, current_user: dict):
    """Background Shopify sync with progress tracking."""
    async with shopify_sync_lock:
        shop_domain, access_token, api_version = get_shopify_config()
        if not shop_domain or not access_token:
            await db.shopify_sync_state.update_one(
                {},
                {"$set": {"status": "error", "error_message": "Shopify credentials not configured"}},
                upsert=True
            )
            await log_system_event("error", "Shopify credentials not configured", {"mode": mode})
            return

        started_at = datetime.utcnow()
        await db.shopify_sync_state.update_one(
            {},
            {"$set": {
                "status": "syncing",
                "mode": mode,
                "products_synced": 0,
                "products_created": 0,
                "products_updated": 0,
                "last_started_at": started_at,
                "last_progress_at": started_at,
                "error_message": None
            }},
            upsert=True
        )
        await log_system_event("info", "Shopify sync started", {"mode": mode, "user": current_user.get("username")})

        try:
            headers = {
                "X-Shopify-Access-Token": access_token,
                "Content-Type": "application/json"
            }

            products_synced = 0
            products_updated = 0
            products_created = 0
            existing_ids = set()
            if mode == "new":
                async for product in db.products.find({"shopify_product_id": {"$exists": True}}, {"shopify_product_id": 1}):
                    existing_ids.add(product["shopify_product_id"])

            next_page_url = f"https://{shop_domain}/admin/api/{api_version}/products.json?limit=50&status=active"

            async with httpx.AsyncClient(timeout=120.0) as client:
                while next_page_url:
                    response = await client.get(next_page_url, headers=headers)
                    if response.status_code != 200:
                        raise RuntimeError(f"Shopify API error: {response.text}")

                    data = response.json()
                    products = data.get("products", [])

                    async def fetch_variant_upc_backup_map(product_id: str) -> dict[str, str]:
                        """Fetch UPC backup metafield for all variants in a product via GraphQL (batched)."""
                        gql_url = f"https://{shop_domain}/admin/api/{api_version}/graphql.json"
                        query = """
                        query($id: ID!, $cursor: String, $ns: String!, $key: String!) {
                          product(id: $id) {
                            variants(first: 100, after: $cursor) {
                              edges {
                                cursor
                                node {
                                  id
                                  metafield(namespace: $ns, key: $key) { value }
                                }
                              }
                              pageInfo { hasNextPage }
                            }
                          }
                        }
                        """
                        out: dict[str, str] = {}
                        cursor = None
                        for attempt in range(6):
                            try:
                                while True:
                                    variables = {
                                        "id": f"gid://shopify/Product/{product_id}",
                                        "cursor": cursor,
                                        "ns": UPC_BACKUP_NAMESPACE,
                                        "key": UPC_BACKUP_KEY,
                                    }
                                    gresp = await client.post(gql_url, headers=headers, json={"query": query, "variables": variables})
                                    if gresp.status_code in (429, 500, 502, 503, 504):
                                        await asyncio.sleep(1.2 * (2 ** attempt))
                                        continue
                                    data = gresp.json() if gresp.status_code == 200 else None
                                    if not data or data.get("errors"):
                                        # throttle or other error
                                        await asyncio.sleep(1.2 * (2 ** attempt))
                                        if attempt >= 5:
                                            return out
                                        break
                                    product = (data.get("data") or {}).get("product") or {}
                                    variants = (product.get("variants") or {})
                                    edges = variants.get("edges") or []
                                    for edge in edges:
                                        node = edge.get("node") or {}
                                        gid = node.get("id") or ""
                                        # gid://shopify/ProductVariant/{id}
                                        m = re.search(r"/ProductVariant/(\\d+)$", gid)
                                        if not m:
                                            continue
                                        vid = m.group(1)
                                        mf = node.get("metafield") or {}
                                        val = mf.get("value")
                                        if val is not None:
                                            digits = re.sub(r"\\D+", "", str(val))
                                            out[vid] = digits or str(val).strip()
                                    if not variants.get("pageInfo", {}).get("hasNextPage"):
                                        return out
                                    cursor = edges[-1].get("cursor")
                                    await asyncio.sleep(0.2)
                            except Exception:
                                await asyncio.sleep(1.2 * (2 ** attempt))
                        return out

                    async def fetch_variant_upc_backup_rest(variant_id: str) -> Optional[str]:
                        """Fallback: fetch UPC backup metafield for a single variant via REST."""
                        url = f"https://{shop_domain}/admin/api/{api_version}/variants/{variant_id}/metafields.json"
                        for attempt in range(4):
                            try:
                                r = await client.get(url, headers=headers)
                                if r.status_code in (429, 500, 502, 503, 504):
                                    await asyncio.sleep(2.0 * (2 ** attempt))
                                    continue
                                if r.status_code != 200:
                                    return None
                                mfs = (r.json() or {}).get("metafields") or []
                                for mf in mfs:
                                    if mf.get("namespace") == UPC_BACKUP_NAMESPACE and mf.get("key") == UPC_BACKUP_KEY:
                                        raw = mf.get("value")
                                        if raw is None:
                                            return None
                                        digits = re.sub(r"\D+", "", str(raw))
                                        return digits or str(raw).strip()
                                return None
                            except Exception:
                                await asyncio.sleep(2.0 * (2 ** attempt))
                            finally:
                                # Shopify REST rate limit: keep under ~2 req/sec
                                await asyncio.sleep(0.65)
                        return None

                    for shopify_product in products:
                        shopify_id = str(shopify_product["id"])
                        if mode == "new" and shopify_id in existing_ids:
                            continue
                        tags_raw = shopify_product.get("tags") or ""
                        tags_list = [t.strip() for t in str(tags_raw).split(",") if t.strip()]

                        # Get first image
                        image_url = None
                        image_base64 = None
                        if shopify_product.get("images"):
                            image_url = shopify_product["images"][0].get("src")
                            if image_url:
                                image_base64 = await download_image_as_base64(image_url)

                        # Create variants + fetch UPC backup metafields (batched GraphQL)
                        upc_backup_map = await fetch_variant_upc_backup_map(shopify_id)

                        # Create variants
                        variants = []
                        for shopify_variant in shopify_product.get("variants", []):
                            v_id = str(shopify_variant.get("id"))
                            upc_backup = upc_backup_map.get(v_id)
                            if not upc_backup:
                                upc_backup = await fetch_variant_upc_backup_rest(v_id)
                            variants.append({
                              "id": str(uuid.uuid4()),
                              "shopify_variant_id": v_id,
                              "product_id": "",  # Will be set below
                              "title": shopify_variant.get("title", "Default"),
                              "sku": shopify_variant.get("sku"),
                              "barcode": shopify_variant.get("barcode"),
                              "upc_backup": upc_backup,
                              "price": float(shopify_variant.get("price", 0)) if shopify_variant.get("price") else 0,
                              "inventory_item_id": str(shopify_variant.get("inventory_item_id")) if shopify_variant.get("inventory_item_id") else None,
                              "inventory_management": shopify_variant.get("inventory_management"),
                              "inventory_quantity": shopify_variant.get("inventory_quantity"),
                              "created_at": datetime.utcnow()
                              })

                        # Check if product exists
                        existing = await db.products.find_one({"shopify_product_id": shopify_id})

                        if existing and mode != "new":
                            # Update existing product - KEEP existing variant IDs for inventory tracking
                            existing_variants = {v.get("shopify_variant_id"): v for v in existing.get("variants", [])}

                            updated_variants = []
                            for new_v in variants:
                                shopify_vid = new_v["shopify_variant_id"]
                                if shopify_vid in existing_variants:
                                    # Keep existing ID, update other fields
                                    old_v = existing_variants[shopify_vid]
                                    new_v["id"] = old_v["id"]
                                    new_v["product_id"] = existing["id"]
                                    if not new_v.get("upc_backup") and old_v.get("upc_backup"):
                                        new_v["upc_backup"] = old_v.get("upc_backup")
                                else:
                                    new_v["product_id"] = existing["id"]
                                updated_variants.append(new_v)

                            await db.products.update_one(
                                {"id": existing["id"]},
                                {"$set": {
                                    "title": shopify_product["title"],
                                    "handle": shopify_product.get("handle"),
                                    "tags": tags_list,
                                    "image_url": image_url,
                                    "image_base64": image_base64,
                                    "variants": updated_variants,
                                    "updated_at": datetime.utcnow()
                                }}
                            )
                            products_updated += 1
                        elif not existing:
                            # Create new product
                            product_id = str(uuid.uuid4())
                            for v in variants:
                                v["product_id"] = product_id

                            product = {
                                "id": product_id,
                                "shopify_product_id": shopify_id,
                                "title": shopify_product["title"],
                                "handle": shopify_product.get("handle"),
                                "tags": tags_list,
                                "image_url": image_url,
                                "image_base64": image_base64,
                                "variants": variants,
                                "created_at": datetime.utcnow(),
                                "updated_at": datetime.utcnow(),
                                "is_active": True
                            }
                            await db.products.insert_one(product)
                            existing_ids.add(shopify_id)
                            products_created += 1

                        products_synced += 1
                        # Throttle between products to avoid Shopify API rate limits
                        await asyncio.sleep(0.4)

                    await db.shopify_sync_state.update_one(
                        {},
                        {"$set": {
                            "products_synced": products_synced,
                            "products_created": products_created,
                            "products_updated": products_updated,
                            "last_progress_at": datetime.utcnow()
                        }},
                        upsert=True
                    )

                    # Check for next page
                    next_page_url = None
                    link_header = response.headers.get("Link", "")
                    if 'rel="next"' in link_header:
                        for part in link_header.split(","):
                            if 'rel="next"' in part:
                                next_page_url = part.split(";")[0].strip().strip("<>")
                                break

            completed_at = datetime.utcnow()
            await db.shopify_sync_state.update_one(
                {},
                {"$set": {
                    "status": "completed",
                    "products_synced": products_synced,
                    "products_created": products_created,
                    "products_updated": products_updated,
                    "last_sync_at": completed_at,
                    "last_completed_at": completed_at,
                    "error_message": None
                }},
                upsert=True
            )
            await log_system_event(
                "info",
                "Shopify sync completed",
                {"mode": mode, "synced": products_synced, "created": products_created, "updated": products_updated}
            )

        except Exception as e:
            logger.error(f"Shopify sync error: {e}")
            await db.shopify_sync_state.update_one(
                {},
                {"$set": {"status": "error", "error_message": str(e), "last_completed_at": datetime.utcnow()}},
                upsert=True
            )
            await log_system_event("error", "Shopify sync failed", {"mode": mode, "error": str(e)})

@api_router.post("/shopify/sync-new")
async def sync_new_shopify_products(current_user: dict = Depends(require_admin)):
    """SINCRONIZZA SOLO NUOVI - Avvia sync in background (READ ONLY)"""
    if shopify_sync_lock.locked():
        return {"status": "syncing", "message": "Sincronizzazione giÃ  in corso"}
    asyncio.create_task(run_shopify_sync("new", current_user))
    return {"status": "syncing", "message": "Sincronizzazione avviata (solo nuovi)"}

@api_router.get("/shopify/sync-status")
async def get_sync_status(current_user: dict = Depends(get_current_user)):
    """Get Shopify sync status"""
    state = await db.shopify_sync_state.find_one({})
    if not state:
        return {
            "status": "never_synced",
            "products_synced": 0,
            "products_created": 0,
            "products_updated": 0,
            "last_sync_at": None,
            "last_started_at": None,
            "last_progress_at": None,
            "last_completed_at": None,
            "mode": None,
            "error_message": None
        }
    if state.get("status") == "syncing":
        last_progress = state.get("last_progress_at")
        if last_progress and isinstance(last_progress, datetime):
            if datetime.utcnow() - last_progress > timedelta(minutes=20):
                await db.shopify_sync_state.update_one(
                    {},
                    {"$set": {
                        "status": "error",
                        "error_message": "Sync bloccata da oltre 20 minuti. Riprova la sincronizzazione.",
                        "last_completed_at": datetime.utcnow()
                    }},
                    upsert=True
                )
                state = await db.shopify_sync_state.find_one({})
    return {
        "status": state.get("status", "idle"),
        "products_synced": state.get("products_synced", 0),
        "products_created": state.get("products_created", 0),
        "products_updated": state.get("products_updated", 0),
        "last_sync_at": state.get("last_sync_at"),
        "last_started_at": state.get("last_started_at"),
        "last_progress_at": state.get("last_progress_at"),
        "last_completed_at": state.get("last_completed_at"),
        "mode": state.get("mode"),
        "error_message": state.get("error_message")
    }

@api_router.post("/shopify/sync-collections")
async def sync_shopify_collections(current_user: dict = Depends(require_admin)):
    """Sync collections from Shopify"""
    shop_domain, access_token, _ = get_shopify_config()
    
    if not shop_domain or not access_token:
        raise HTTPException(status_code=400, detail="Shopify credentials not configured")
    
    try:
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        collections_synced = 0
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Sync custom collections
            custom_collections_url = f"https://{shop_domain}/admin/api/2024-01/custom_collections.json?limit=250"
            response = await client.get(custom_collections_url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                for shopify_coll in data.get("custom_collections", []):
                    # Get products in collection
                    collects_url = f"https://{shop_domain}/admin/api/2024-01/collects.json?collection_id={shopify_coll['id']}&limit=250"
                    collects_resp = await client.get(collects_url, headers=headers)
                    
                    product_ids = []
                    if collects_resp.status_code == 200:
                        collects_data = collects_resp.json()
                        shopify_product_ids = [str(c["product_id"]) for c in collects_data.get("collects", [])]
                        
                        # Map to our product IDs
                        for spid in shopify_product_ids:
                            product = await db.products.find_one({"shopify_product_id": spid})
                            if product:
                                product_ids.append(product["id"])
                    
                    # Upsert collection
                    existing = await db.collections.find_one({"shopify_collection_id": str(shopify_coll["id"])})
                    if existing:
                        await db.collections.update_one(
                            {"id": existing["id"]},
                            {"$set": {
                                "title": shopify_coll["title"],
                                "handle": shopify_coll.get("handle"),
                                "product_ids": product_ids
                            }}
                        )
                    else:
                        collection = Collection(
                            shopify_collection_id=str(shopify_coll["id"]),
                            title=shopify_coll["title"],
                            handle=shopify_coll.get("handle"),
                            product_ids=product_ids
                        )
                        await db.collections.insert_one(collection.dict())
                    
                    collections_synced += 1
            
            # Sync smart collections
            smart_collections_url = f"https://{shop_domain}/admin/api/2024-01/smart_collections.json?limit=250"
            response = await client.get(smart_collections_url, headers=headers)
            
            if response.status_code == 200:
                data = response.json()
                for shopify_coll in data.get("smart_collections", []):
                    # Get products in collection
                    products_url = f"https://{shop_domain}/admin/api/2024-01/collections/{shopify_coll['id']}/products.json?limit=250"
                    products_resp = await client.get(products_url, headers=headers)
                    
                    product_ids = []
                    if products_resp.status_code == 200:
                        products_data = products_resp.json()
                        for p in products_data.get("products", []):
                            product = await db.products.find_one({"shopify_product_id": str(p["id"])})
                            if product:
                                product_ids.append(product["id"])
                    
                    # Upsert collection
                    existing = await db.collections.find_one({"shopify_collection_id": str(shopify_coll["id"])})
                    if existing:
                        await db.collections.update_one(
                            {"id": existing["id"]},
                            {"$set": {
                                "title": shopify_coll["title"],
                                "handle": shopify_coll.get("handle"),
                                "product_ids": product_ids
                            }}
                        )
                    else:
                        collection = Collection(
                            shopify_collection_id=str(shopify_coll["id"]),
                            title=shopify_coll["title"],
                            handle=shopify_coll.get("handle"),
                            product_ids=product_ids
                        )
                        await db.collections.insert_one(collection.dict())
                    
                    collections_synced += 1
        
        return {"message": f"Synced {collections_synced} collections from Shopify"}
        
    except Exception as e:
        logger.error(f"Shopify collections sync error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== SHOPIFY INVENTORY SYNC (Background) ====================

# Global variable to track running sync task
inventory_sync_task = None

async def background_inventory_sync(user_id: str, username: str):
    """Background task to sync inventory from Shopify with rate limiting"""
    shop_domain, access_token, api_version = get_shopify_config()
    if not shop_domain or not access_token:
        await db.inventory_sync_status.update_one(
            {},
            {"$set": {"status": "error", "error_message": "Shopify credentials not configured"}},
            upsert=True
        )
        await log_system_event("error", "Inventory sync failed: missing Shopify credentials")
        return
    
    try:
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        # Update status: starting
        await db.inventory_sync_status.update_one(
            {},
            {"$set": {
                "status": "running",
                "started_at": datetime.utcnow(),
                "started_by": username,
                "progress": 0,
                "total_variants": 0,
                "processed_variants": 0,
                "imported_count": 0,
                "error_message": None,
                "estimated_time_remaining": None
            }},
            upsert=True
        )
        
        # Get Shopify locations
        async with httpx.AsyncClient(timeout=60.0) as client:
            locations_url = f"https://{shop_domain}/admin/api/{api_version}/locations.json"
            loc_response = await client.get(locations_url, headers=headers)
            
            if loc_response.status_code != 200:
                raise Exception("Failed to get Shopify locations")
            
            shopify_locations = loc_response.json().get("locations", [])
            
            warehouse_location_id = None
            for loc in shopify_locations:
                if "warehouse" in loc.get("name", "").lower():
                    warehouse_location_id = loc["id"]
                    break
            
            if not warehouse_location_id and shopify_locations:
                warehouse_location_id = shopify_locations[0]["id"]
            
            if not warehouse_location_id:
                raise Exception("No Shopify location found")
        
        # Get local location
        local_location = await db.locations.find_one({"name": {"$regex": "warehouse", "$options": "i"}})
        if not local_location:
            local_location = await db.locations.find_one({"is_active": True})
        if not local_location:
            local_location = Location(name="Warehouse", description="Magazzino principale da Shopify")
            await db.locations.insert_one(local_location.dict())
            local_location = local_location.dict()
        
        local_location_id = local_location.get("id") or local_location["id"]
        
        # Fetch products and prepare variants (optimized)
        products = await db.products.find({"is_active": True, "shopify_product_id": {"$ne": None}}, {"_id": 0}).to_list(10000)
        candidates = []
        for p in products:
            for v in p.get("variants", []):
                inv_item_id = v.get("inventory_item_id")
                if not inv_item_id:
                    continue
                candidates.append({
                    "variant_id": v.get("id"),
                    "inventory_item_id": str(inv_item_id),
                    "product_id": p.get("id")
                })

        total_variants = len(candidates)

        if total_variants == 0:
            await db.inventory_sync_status.update_one(
                {},
                {"$set": {
                    "status": "error",
                    "error_message": "Nessuna variante con inventory tracciato. Sincronizza prima i prodotti da Shopify o abilita il tracking inventario.",
                    "completed_at": datetime.utcnow()
                }},
                upsert=True
            )
            await log_system_event("warning", "Inventory sync skipped: no tracked variants")
            return

        # Delete only Shopify-sourced inventory to preserve local stock
        await db.inventory_levels.delete_many({"source": "shopify"})
        
        await db.inventory_sync_status.update_one(
            {},
            {"$set": {"total_variants": total_variants}}
        )
        
        imported_count = 0
        processed_count = 0
        start_time = datetime.utcnow()
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Batch inventory levels by inventory_item_id
            chunk_size = 50
            for i in range(0, len(candidates), chunk_size):
                chunk = candidates[i:i + chunk_size]
                inv_item_ids = ",".join([c["inventory_item_id"] for c in chunk])
                inv_url = f"https://{shop_domain}/admin/api/{api_version}/inventory_levels.json?inventory_item_ids={inv_item_ids}&location_ids={warehouse_location_id}"
                try:
                    inv_response = await client.get(inv_url, headers=headers)
                    if inv_response.status_code == 429:
                        await asyncio.sleep(2)
                        inv_response = await client.get(inv_url, headers=headers)

                    if inv_response.status_code == 200:
                        inv_levels = inv_response.json().get("inventory_levels", [])
                        available_map = {str(x.get("inventory_item_id")): x.get("available", 0) for x in inv_levels}
                        for c in chunk:
                            available = available_map.get(c["inventory_item_id"], 0)
                            if available and available > 0:
                                new_level = InventoryLevel(
                                    variant_id=c["variant_id"],
                                    product_id=c.get("product_id"),
                                    location_id=local_location_id,
                                    shelf_id=None,
                                    quantity=available,
                                    source="shopify",
                                    last_synced_at=datetime.utcnow()
                                )
                                await db.inventory_levels.insert_one(new_level.dict())
                                imported_count += 1

                    processed_count += len(chunk)

                    progress = int((processed_count / total_variants) * 100) if total_variants > 0 else 0
                    elapsed = (datetime.utcnow() - start_time).total_seconds()
                    if processed_count > 0:
                        rate = processed_count / elapsed
                        remaining = total_variants - processed_count
                        eta_seconds = int(remaining / rate) if rate > 0 else 0
                        eta_minutes = eta_seconds // 60
                        eta_str = f"{eta_minutes} minuti" if eta_minutes > 0 else f"{eta_seconds} secondi"
                    else:
                        eta_str = "Calcolo in corso..."

                    await db.inventory_sync_status.update_one(
                        {},
                        {"$set": {
                            "progress": progress,
                            "processed_variants": processed_count,
                            "imported_count": imported_count,
                            "estimated_time_remaining": eta_str
                        }}
                    )

                    await asyncio.sleep(0.2)

                except Exception as e:
                    logger.error(f"Error processing inventory chunk: {e}")
                    processed_count += len(chunk)
                    continue
        
        # Sync completed
        await db.inventory_sync_status.update_one(
            {},
            {"$set": {
                "status": "completed",
                "completed_at": datetime.utcnow(),
                "progress": 100,
                "processed_variants": processed_count,
                "imported_count": imported_count,
                "estimated_time_remaining": None
            }}
        )
        
        logger.info(f"Inventory sync completed: {imported_count} items imported")
        
    except Exception as e:
        logger.error(f"Inventory sync error: {e}")
        await db.inventory_sync_status.update_one(
            {},
            {"$set": {
                "status": "error",
                "error_message": str(e),
                "completed_at": datetime.utcnow()
            }}
        )

@api_router.post("/shopify/sync-inventory")
async def start_inventory_sync(background_tasks: BackgroundTasks, current_user: dict = Depends(require_admin)):
    """Start background inventory sync from Shopify"""
    global inventory_sync_task
    
    # Check if sync is already running
    status = await db.inventory_sync_status.find_one({}, {"_id": 0})
    if status and status.get("status") == "running":
        return {
            "message": "Sincronizzazione già in corso",
            "status": "running",
            "progress": status.get("progress", 0)
        }
    
    shop_domain, access_token, _ = get_shopify_config()
    
    if not shop_domain or not access_token:
        raise HTTPException(status_code=400, detail="Credenziali Shopify non configurate")
    
    # Start background task
    asyncio.create_task(background_inventory_sync(current_user["id"], current_user["username"]))
    
    return {
        "message": "Sincronizzazione inventario avviata in background",
        "status": "started",
        "note": "La sincronizzazione continuerà anche se chiudi l'app. Controlla lo stato nella pagina impostazioni."
    }

@api_router.get("/shopify/inventory-sync-status")
async def get_inventory_sync_status(current_user: dict = Depends(get_current_user)):
    """Get current inventory sync status"""
    status = await db.inventory_sync_status.find_one({}, {"_id": 0})
    if not status:
        return {
            "status": "never_started",
            "progress": 0,
            "message": "Sincronizzazione mai avviata"
        }
    return status

@api_router.post("/shopify/sync-inventory/stop")
async def stop_inventory_sync(current_user: dict = Depends(require_admin)):
    """Stop running inventory sync"""
    await db.inventory_sync_status.update_one(
        {},
        {"$set": {"status": "stopped", "completed_at": datetime.utcnow()}}
    )
    return {"message": "Sincronizzazione interrotta"}

@api_router.get("/shopify/orders/pending")
async def get_pending_shopify_orders(current_user: dict = Depends(get_current_user)):
    """Get unfulfilled orders from Shopify (orders to ship)"""
    shop_domain, access_token, api_version = get_shopify_config()
    
    if not shop_domain or not access_token:
        raise HTTPException(status_code=400, detail="Shopify credentials not configured")
    
    try:
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Get unfulfilled orders
            orders_url = f"https://{shop_domain}/admin/api/{api_version}/orders.json?status=unfulfilled&limit=50"
            response = await client.get(orders_url, headers=headers)
            
            if response.status_code != 200:
                raise HTTPException(status_code=500, detail="Failed to get Shopify orders")
            
            orders = response.json().get("orders", [])
            
            # Format orders for frontend
            pending_orders = []
            for order in orders:
                order_items = []
                for item in order.get("line_items", []):
                    # Find our local product/variant
                    product = await db.products.find_one(
                        {"shopify_product_id": str(item.get("product_id"))},
                        {"_id": 0}
                    )
                    
                    variant_info = None
                    if product:
                        for v in product.get("variants", []):
                            if v.get("shopify_variant_id") == str(item.get("variant_id")):
                                variant_info = v
                                break
                    
                    order_items.append({
                        "line_item_id": item.get("id"),
                        "product_id": str(item.get("product_id")),
                        "variant_id": str(item.get("variant_id")),
                        "title": item.get("title"),
                        "variant_title": item.get("variant_title"),
                        "quantity": item.get("quantity"),
                        "fulfillable_quantity": item.get("fulfillable_quantity"),
                        "price": item.get("price"),
                        "sku": item.get("sku"),
                        "local_product_id": product.get("id") if product else None,
                        "local_variant_id": variant_info.get("id") if variant_info else None,
                        "image_base64": product.get("image_base64") if product else None
                    })
                
                pending_orders.append({
                    "id": order.get("id"),
                    "order_number": order.get("order_number"),
                    "name": order.get("name"),  # e.g., "#1001"
                    "created_at": order.get("created_at"),
                    "total_price": order.get("total_price"),
                    "currency": order.get("currency"),
                    "fulfillment_status": order.get("fulfillment_status"),
                    "customer": {
                        "name": f"{order.get('customer', {}).get('first_name', '')} {order.get('customer', {}).get('last_name', '')}".strip(),
                        "email": order.get("customer", {}).get("email")
                    },
                    "shipping_address": order.get("shipping_address", {}),
                    "items": order_items,
                    "items_count": len(order_items)
                })
            
            return {"orders": pending_orders, "total": len(pending_orders)}
        
    except Exception as e:
        logger.error(f"Error getting Shopify orders: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/shopify/orders/{order_id}/mark-shipped")
async def mark_order_shipped(order_id: str, current_user: dict = Depends(get_current_user)):
    """Mark order as shipped locally (removes from pending list)"""
    # For now, we just track shipped orders locally
    # In future with webhook, this will also update Shopify
    
    shipped_order = {
        "id": str(uuid.uuid4()),
        "shopify_order_id": order_id,
        "shipped_by": current_user["id"],
        "shipped_by_username": current_user["username"],
        "shipped_at": datetime.utcnow()
    }
    
    await db.shipped_orders.insert_one(shipped_order)
    
    await log_action(
        ActionType.SHOPIFY_UPDATE,
        current_user,
        f"Marked Shopify order #{order_id} as shipped",
        entity_type="shopify_order",
        entity_id=order_id
    )
    
    return {"message": "Order marked as shipped"}

@api_router.get("/shopify/orders/shipped")
async def get_shipped_orders(skip: int = 0, limit: int = 50, current_user: dict = Depends(get_current_user)):
    """Get locally shipped orders"""
    total = await db.shipped_orders.count_documents({})
    orders = await db.shipped_orders.find({}, {"_id": 0}).sort("shipped_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"orders": orders, "total": total}

# ==================== SHOPIFY INVENTORY UPDATE (when selling locally) ====================

async def update_shopify_inventory(variant_id: str, quantity_change: int, current_user: dict):
    """Update Shopify inventory when selling locally"""
    shop_domain, access_token, api_version = get_shopify_config()
    
    if not shop_domain or not access_token:
        logger.warning("Shopify credentials not configured, skipping inventory update")
        await log_system_event("warning", "Skipping Shopify inventory update (missing credentials)")
        return False
    
    try:
        # Get variant's Shopify variant ID
        product = await db.products.find_one({"variants.id": variant_id})
        if not product:
            return False
        product_tags = [str(t).strip().upper() for t in (product.get("tags") or []) if str(t).strip()]
        if "NOGESTIONALE" in product_tags:
            logger.info(f"Skipping Shopify sync for variant {variant_id}: product tagged NOGESTIONALE")
            return False
        
        shopify_variant_id = None
        for v in product.get("variants", []):
            if v.get("id") == variant_id:
                shopify_variant_id = v.get("shopify_variant_id")
                break
        
        if not shopify_variant_id:
            logger.warning(f"No Shopify variant ID for variant {variant_id}")
            return False
        
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Get inventory item ID
            variant_url = f"https://{shop_domain}/admin/api/{api_version}/variants/{shopify_variant_id}.json"
            var_response = await client.get(variant_url, headers=headers)
            
            if var_response.status_code != 200:
                logger.error(f"Failed to get variant: {var_response.status_code}")
                return False
            
            inventory_item_id = var_response.json().get("variant", {}).get("inventory_item_id")
            
            if not inventory_item_id:
                return False
            
            # Get locations
            locations_url = f"https://{shop_domain}/admin/api/{api_version}/locations.json"
            loc_response = await client.get(locations_url, headers=headers)
            
            if loc_response.status_code != 200:
                return False
            
            shopify_locations = loc_response.json().get("locations", [])
            warehouse_location_id = None
            
            for loc in shopify_locations:
                if "warehouse" in loc.get("name", "").lower():
                    warehouse_location_id = loc["id"]
                    break
            
            if not warehouse_location_id and shopify_locations:
                warehouse_location_id = shopify_locations[0]["id"]
            
            if not warehouse_location_id:
                return False
            
            # Adjust inventory on Shopify
            adjust_url = f"https://{shop_domain}/admin/api/{api_version}/inventory_levels/adjust.json"
            adjust_data = {
                "location_id": warehouse_location_id,
                "inventory_item_id": inventory_item_id,
                "available_adjustment": quantity_change  # negative for sales
            }
            
            adjust_response = await client.post(adjust_url, headers=headers, json=adjust_data)
            
            if adjust_response.status_code == 200:
                logger.info(f"Updated Shopify inventory for variant {shopify_variant_id}: {quantity_change}")
                
                await log_action(
                    ActionType.SHOPIFY_UPDATE,
                    current_user,
                    f"Updated Shopify inventory: {quantity_change} for {product.get('title')}",
                    entity_type="shopify_inventory",
                    metadata={"shopify_variant_id": shopify_variant_id, "adjustment": quantity_change}
                )
                
                return True
            else:
                logger.error(f"Failed to adjust Shopify inventory: {adjust_response.text}")
                return False
                
    except Exception as e:
        logger.error(f"Error updating Shopify inventory: {e}")
        return False

# ==================== LOCAL PRODUCTS (non-Shopify) ====================

@api_router.post("/products/local")
async def create_local_product(data: LocalProductCreate, current_user: dict = Depends(get_current_user)):
    """Create a local product (not synced to Shopify)"""
    # If barcode exists on Shopify product, return that instead (avoid duplicates)
    barcodes = []
    if data.variants:
        for v in data.variants:
            if v.barcode:
                barcodes.append(v.barcode.strip())
    else:
        if data.barcode:
            barcodes.append(data.barcode.strip())

    if barcodes:
        existing_shopify = await db.products.find_one(
            {
                "shopify_product_id": {"$ne": None},
                "variants.barcode": {"$in": barcodes}
            },
            {"_id": 0, "id": 1, "title": 1, "variants": 1}
        )
        if existing_shopify:
            matched = None
            for v in existing_shopify.get("variants", []):
                if v.get("barcode") in barcodes:
                    matched = v.get("barcode")
                    break
            return {
                "message": "Barcode giÃ  presente su Shopify, prodotto collegato",
                "matched_shopify_product": True,
                "product_id": existing_shopify.get("id"),
                "product_title": existing_shopify.get("title"),
                "matched_barcode": matched
            }
    
    # Create product
    product = Product(
        title=data.title,
        shopify_product_id=None,  # Local product, no Shopify ID
        handle=data.title.lower().replace(" ", "-"),
        image_base64=data.image_base64,
        variants=[]
    )
    
    # Create variants or single variant
    if data.variants:
        for var_data in data.variants:
            variant = ProductVariant(
                product_id=product.id,
                shopify_variant_id=None,
                title=var_data.title,
                sku=var_data.sku,
                barcode=var_data.barcode,
                price=var_data.price
            )
            product.variants.append(variant)
    else:
        # Single product without sizes
        variant = ProductVariant(
            product_id=product.id,
            shopify_variant_id=None,
            title="Default",
            sku=data.sku,
            barcode=data.barcode,
            price=data.price
        )
        product.variants.append(variant)
    
    await db.products.insert_one(product.dict())
    
    await log_action(
        ActionType.SHOPIFY_SYNC,  # Reusing, could create new action type
        current_user,
        f"Created local product: {data.title}",
        entity_type="product",
        entity_id=product.id,
        new_data={"title": data.title, "price": data.price, "variants_count": len(product.variants)}
    )
    
    return {
        "id": product.id,
        "title": product.title,
        "variants": len(product.variants),
        "product": product.dict(),
        "matched_shopify_product": False
    }

@api_router.get("/products/local")
async def get_local_products(current_user: dict = Depends(get_current_user)):
    """Get all local products (not from Shopify)"""
    products = await db.products.find(
        {"shopify_product_id": None, "is_active": True},
        {"_id": 0}
    ).to_list(1000)
    return products

# ==================== PURCHASE LINKS (Acquisto da Fornitore) ====================

@api_router.post("/purchase-links")
async def create_purchase_link(data: PurchaseLinkCreate, current_user: dict = Depends(get_current_user)):
    """Create a purchase link to send to supplier"""
    doc_type = (data.doc_type or "acquisto").strip().lower()
    if doc_type not in ("acquisto", "contovendita"):
        raise HTTPException(status_code=400, detail="Tipo documento non valido")
    
    # Calculate total
    total = 0.0 if doc_type == "contovendita" else sum(item.purchase_price * item.quantity for item in data.items)
    
    # Create link with 2 days expiration
    purchase_link = PurchaseLink(
        items=[item.dict() for item in data.items],
        total_amount=total,
        note=data.note,
        created_by=current_user["id"],
        created_by_username=current_user["username"],
        expires_at=datetime.utcnow() + timedelta(days=2),
        doc_type=doc_type
    )
    
    await db.purchase_links.insert_one(purchase_link.dict())
    
    await log_action(
        ActionType.SHOPIFY_UPDATE,  # Reusing
        current_user,
        f"Created purchase link: EUR {total:.2f} for {len(data.items)} items",
        entity_type="purchase_link",
        entity_id=purchase_link.id,
        new_data={"total": total, "items_count": len(data.items), "doc_type": doc_type}
    )
    
    return {
        "id": purchase_link.id,
        "token": purchase_link.token,
        "total_amount": total,
        "expires_at": purchase_link.expires_at,
        "items_count": len(data.items),
        "doc_type": doc_type
    }

@api_router.get("/purchase-links")
async def get_purchase_links(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all purchase links (admin view)"""
    query = {}
    if status:
        query["status"] = status
    
    links = await db.purchase_links.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Check and update expired links
    now = datetime.utcnow()
    for link in links:
        if link.get("status") == "pending" and link.get("expires_at"):
            expires = link["expires_at"]
            if isinstance(expires, str):
                expires = datetime.fromisoformat(expires.replace("Z", "+00:00"))
            if expires < now:
                await db.purchase_links.update_one(
                    {"id": link["id"]},
                    {"$set": {"status": "expired"}}
                )
                link["status"] = "expired"
    
    return links

@api_router.get("/purchase-links/{link_id}")
async def get_purchase_link(link_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific purchase link"""
    link = await db.purchase_links.find_one({"id": link_id}, {"_id": 0})
    if not link:
        raise HTTPException(status_code=404, detail="Purchase link not found")
    return link

@api_router.delete("/purchase-links/{link_id}")
async def delete_purchase_link(link_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a purchase link"""
    result = await db.purchase_links.delete_one({"id": link_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase link not found")
    return {"message": "Purchase link deleted"}

# PUBLIC ENDPOINTS (no auth required) for suppliers

@api_router.get("/public/purchase/{token}")
async def get_public_purchase_link(token: str):
    """Public endpoint for supplier to view purchase details"""
    link = await db.purchase_links.find_one({"token": token}, {"_id": 0})
    
    if not link:
        raise HTTPException(status_code=404, detail="Link non trovato")
    
    # Check expiration
    expires = link.get("expires_at")
    if expires:
        if isinstance(expires, str):
            expires = datetime.fromisoformat(expires.replace("Z", "+00:00"))
        if expires < datetime.utcnow():
            return {"error": "Link scaduto", "expired": True}
    
    if link.get("status") == "submitted":
        return {"error": "Questo link è già stato compilato", "already_submitted": True}
    
    # Enrich items with product image when available
    enriched_items = []
    for item in link.get("items", []):
        enriched = dict(item)
        product_id = item.get("product_id")
        if product_id:
            product = await db.products.find_one({"id": product_id}, {"_id": 0})
            if product:
                enriched["product_image"] = product.get("image_base64") or product.get("image_url")
        enriched_items.append(enriched)

    # Return safe data for public view
    return {
        "id": link["id"],
        "items": enriched_items,
        "total_amount": link["total_amount"],
        "note": link.get("note"),
        "created_at": link["created_at"],
        "expires_at": link["expires_at"],
        "status": link["status"],
        "doc_type": link.get("doc_type", "acquisto")
    }

@api_router.post("/public/purchase/{token}/submit")
async def submit_supplier_data(token: str, supplier_data: SupplierData):
    """Public endpoint for supplier to submit their data"""
    link = await db.purchase_links.find_one({"token": token})
    
    if not link:
        raise HTTPException(status_code=404, detail="Link non trovato")
    
    # Check expiration
    expires = link.get("expires_at")
    if expires:
        if isinstance(expires, str):
            expires = datetime.fromisoformat(expires.replace("Z", "+00:00"))
        if expires < datetime.utcnow():
            raise HTTPException(status_code=400, detail="Link scaduto")
    
    if link.get("status") == "submitted":
        raise HTTPException(status_code=400, detail="Già compilato")
    
    # Validate required fields by document type
    doc_type = (link.get("doc_type") or "acquisto").strip().lower()
    required_fields = {
        "first_name": supplier_data.first_name,
        "last_name": supplier_data.last_name,
        "birth_date": supplier_data.birth_date,
        "birth_place": supplier_data.birth_place,
        "birth_country": supplier_data.birth_country,
        "residence_address": supplier_data.residence_address,
        "residence_city": supplier_data.residence_city,
        "residence_province": supplier_data.residence_province,
        "residence_cap": supplier_data.residence_cap,
        "residence_country": supplier_data.residence_country,
        "fiscal_code": supplier_data.fiscal_code,
        "phone": supplier_data.phone,
        "signature": supplier_data.signature,
    }
    if doc_type == "acquisto":
        required_fields["iban"] = supplier_data.iban

    missing = [k for k, v in required_fields.items() if not (v and str(v).strip())]
    if missing:
        raise HTTPException(status_code=400, detail={"message": "Campi obbligatori mancanti", "fields": missing})

    # Update link with supplier data
    await db.purchase_links.update_one(
        {"token": token},
        {
            "$set": {
                "status": "submitted",
                "supplier_data": supplier_data.dict(exclude_none=True),
                "submitted_at": datetime.utcnow()
            }
        }
    )
    
    return {"message": "Dati inviati con successo. Grazie!"}

@api_router.get("/purchase-links/{link_id}/pdf")
async def generate_purchase_pdf(
    link_id: str,
    token: Optional[str] = None,
    current_user: dict = Depends(get_current_user_or_token)
):
    """Generate legal one-page PDF for private purchase or contovendita."""
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader

    link = await db.purchase_links.find_one({"id": link_id})
    if not link:
        raise HTTPException(status_code=404, detail="Link non trovato")

    if not link.get("supplier_data"):
        raise HTTPException(status_code=400, detail="Dati fornitore non ancora compilati")

    supplier = link["supplier_data"]

    doc_type = (link.get("doc_type") or "acquisto").strip().lower()
    if doc_type not in ("acquisto", "contovendita"):
        doc_type = "acquisto"

    if doc_type == "contovendita":
        template_path = ROOT_DIR / "CONTOVENDITAFOGLIO.pdf"
        if not template_path.exists():
            raise HTTPException(status_code=500, detail="Template CONTOVENDITAFOGLIO.pdf mancante")
    else:
        template_path = ROOT_DIR / "ACQUISTODAPRIVATO.pdf"
        if not template_path.exists():
            raise HTTPException(status_code=500, detail="Template ACQUISTODAPRIVATO.pdf mancante")

    from pypdf import PdfReader, PdfWriter
    from copy import deepcopy

    tmpl_reader = PdfReader(str(template_path))
    tmpl_page = tmpl_reader.pages[0]
    page_w = float(tmpl_page.mediabox.width)
    page_h = float(tmpl_page.mediabox.height)

    def y_from_img(row_from_top: float, offset: float = 10.0) -> float:
        return page_h - row_from_top + offset

    def draw_text_fit(c, text: str, x: float, y: float, max_w: float, font: str = "Helvetica", size: float = 24.0):
        t = (text or "").strip()
        if not t:
            return
        c.setFont(font, size)
        width = c.stringWidth(t, font, size)
        if width <= max_w:
            c.drawString(x, y, t)
            return
        # shrink to fit
        min_size = 14.0
        cur = size
        while cur > min_size and c.stringWidth(t, font, cur) > max_w:
            cur -= 0.5
        c.setFont(font, cur)
        c.drawString(x, y, t)

    def parse_signature_paths(signature_value: str):
        svg_xml = re.sub(r"\s+", " ", signature_value)
        width_match = re.search(r'width="([\d\.]+)"', svg_xml)
        height_match = re.search(r'height="([\d\.]+)"', svg_xml)
        view_w = float(width_match.group(1)) if width_match else 600.0
        view_h = float(height_match.group(1)) if height_match else 180.0
        paths = []
        for d in re.findall(r'd="([^"]+)"', svg_xml):
            tokens = re.findall(r'([ML])\s*([-\d\.]+)\s*([-\d\.]+)', d)
            if tokens:
                paths.append([(cmd, float(x), float(y)) for cmd, x, y in tokens])
        return paths, view_w, view_h

    def draw_signature(c, signature_value: str, x: float, y: float, w: float, h: float):
        if not signature_value:
            return
        try:
            if str(signature_value).startswith("data:image/svg+xml"):
                svg_data = signature_value.split(",", 1)[1]
                svg_xml = unquote(svg_data)
                paths, view_w, view_h = parse_signature_paths(svg_xml)
                if not paths or view_w <= 0 or view_h <= 0:
                    return
                scale_x = w / view_w
                scale_y = h / view_h
                c.setLineWidth(1)
                for path in paths:
                    last = None
                    for cmd, px, py in path:
                        dx = x + px * scale_x
                        dy = y + (view_h - py) * scale_y
                        if cmd == 'M':
                            last = (dx, dy)
                        elif cmd == 'L' and last:
                            c.line(last[0], last[1], dx, dy)
                            last = (dx, dy)
            elif str(signature_value).startswith("data:image"):
                _, b64 = signature_value.split(",", 1)
                img_bytes = base64.b64decode(b64)
                c.drawImage(ImageReader(BytesIO(img_bytes)), x, y, width=w, height=h, mask='auto')
        except Exception as e:
            logger.warning(f"Signature render failed: {e}")

    # Build overlay
    overlay = BytesIO()
    c = canvas.Canvas(overlay, pagesize=(page_w, page_h))
    c.setFillColorRGB(0, 0, 0)

    name = f"{supplier.get('first_name', '')} {supplier.get('last_name', '')}".strip()
    items = link.get("items", [])
    product_size = 22.0

    if doc_type == "contovendita":
        # --- Field positions for CONTOVENDITAFOGLIO ---
        draw_text_fit(c, name, 212, y_from_img(140), 809 - 212)
        draw_text_fit(c, supplier.get("birth_place", ""), 72, y_from_img(179), 458 - 72)
        draw_text_fit(c, supplier.get("birth_country", ""), 469, y_from_img(179), 502 - 469)
        draw_text_fit(c, supplier.get("birth_date", ""), 530, y_from_img(179), 722 - 530)

        draw_text_fit(c, supplier.get("residence_address", ""), 72, y_from_img(218), 467 - 72)
        draw_text_fit(c, supplier.get("residence_city", ""), 107, y_from_img(254), 458 - 107)
        draw_text_fit(c, supplier.get("residence_province", ""), 469, y_from_img(254), 502 - 469)
        draw_text_fit(c, supplier.get("residence_cap", ""), 562, y_from_img(254), 666 - 562)

        draw_text_fit(c, supplier.get("phone", ""), 136, y_from_img(299), 328 - 136)
        draw_text_fit(c, supplier.get("email", ""), 398, y_from_img(299), 757 - 398)

        draw_text_fit(c, supplier.get("fiscal_code", ""), 177, y_from_img(341), 774 - 177)

        product_rows = [551, 593, 638, 680, 722, 767, 809, 854, 896, 938, 983, 1025]
        for idx, row in enumerate(product_rows):
            if idx >= len(items):
                break
            item = items[idx]
            title = (item.get("title", "") or "").strip()
            variant = (item.get("variant_title", "") or "").strip()
            qty = int(item.get("quantity", 1) or 1)
            price = float(item.get("purchase_price", 0) or 0)

            brand = ""
            model = title
            if " " in title:
                brand, model = title.split(" ", 1)
            if qty > 1:
                model = f"{model} x{qty}"

            y = y_from_img(row, offset=8)
            draw_text_fit(c, brand, 70, y, 140, size=product_size)
            draw_text_fit(c, model, 210, y, 420, size=product_size)
            draw_text_fit(c, variant or "-", 640, y, 90, size=product_size)
            if price > 0:
                draw_text_fit(c, f"{price:.2f}", 735, y, 90, size=product_size)

        draw_text_fit(c, datetime.utcnow().strftime('%d/%m/%Y'), 165, y_from_img(1367), 339 - 165)

        sig_x = 70
        sig_y = y_from_img(1460) + 6
        sig_w = 228
        sig_h = 55
        draw_signature(c, supplier.get("signature", "") or "", sig_x, sig_y, sig_w, sig_h)
    else:
        # --- Field positions for ACQUISTODAPRIVATO ---
        draw_text_fit(c, name, 238, y_from_img(114), 990 - 238)

        draw_text_fit(c, supplier.get("birth_place", ""), 72, y_from_img(153), 558 - 72)
        draw_text_fit(c, supplier.get("birth_country", ""), 570, y_from_img(153), 614 - 570)
        draw_text_fit(c, supplier.get("birth_date", ""), 646, y_from_img(153), 889 - 646)

        draw_text_fit(c, supplier.get("residence_address", ""), 72, y_from_img(195), 569 - 72)
        draw_text_fit(c, supplier.get("residence_city", ""), 116, y_from_img(237), 558 - 116)
        draw_text_fit(c, supplier.get("residence_province", ""), 571, y_from_img(237), 615 - 571)
        draw_text_fit(c, supplier.get("residence_cap", ""), 695, y_from_img(237), 828 - 695)

        draw_text_fit(c, supplier.get("phone", ""), 151, y_from_img(282), 394 - 151)
        draw_text_fit(c, supplier.get("email", ""), 483, y_from_img(282), 936 - 483)

        draw_text_fit(c, supplier.get("fiscal_code", ""), 231, y_from_img(327), 983 - 231)

        product_rows = [585, 630, 675, 717, 762, 807, 852, 897, 942, 987, 1029, 1074]
        for idx, row in enumerate(product_rows):
            if idx >= len(items):
                break
            item = items[idx]
            title = (item.get("title", "") or "").strip()
            variant = (item.get("variant_title", "") or "").strip()
            qty = int(item.get("quantity", 1) or 1)
            price = float(item.get("purchase_price", 0) or 0)

            brand = ""
            model = title
            if " " in title:
                brand, model = title.split(" ", 1)
            if qty > 1:
                model = f"{model} x{qty}"

            y = y_from_img(row, offset=8)
            draw_text_fit(c, brand, 70, y, 160, size=product_size)
            draw_text_fit(c, model, 230, y, 560, size=product_size)
            draw_text_fit(c, variant or "-", 820, y, 120, size=product_size)
            draw_text_fit(c, f"{price:.2f}", 960, y, 120, size=product_size)

        total_amount = float(link.get("total_amount", 0) or 0)
        draw_text_fit(c, f"{total_amount:.2f}", 465, y_from_img(1170), 686 - 465)

        draw_text_fit(c, datetime.utcnow().strftime('%d/%m/%Y'), 186, y_from_img(1392), 407 - 186)

        sig_x = 70
        sig_y = y_from_img(1482) + 10
        sig_w = 287
        sig_h = 60
        draw_signature(c, supplier.get("signature", "") or "", sig_x, sig_y, sig_w, sig_h)

    c.save()
    overlay.seek(0)

    # Merge overlay on top of template
    overlay_reader = PdfReader(overlay)
    writer = PdfWriter()
    base = deepcopy(tmpl_reader.pages[0])
    base.merge_page(overlay_reader.pages[0])
    writer.add_page(base)

    final_buffer = BytesIO()
    writer.write(final_buffer)
    final_buffer.seek(0)

    filename = f"acquisto_privato_{link_id[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        final_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def _build_inventory_export_rows(
    location_id: Optional[str],
    collection_product_ids: Optional[list],
    search: Optional[str] = None,
    size: Optional[str] = None,
    product_ids: Optional[list] = None,
    variant_ids: Optional[list] = None,
):
    inventory = await _get_inventory_levels_with_fallback(location_id)
    rows = []
    search_l = search.lower().strip() if search else None
    size_l = size.lower().strip() if size else None
    for inv in inventory:
        try:
            qty = int(inv.get("quantity", 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue

        product = await db.products.find_one({"variants.id": inv.get("variant_id")})
        if not product and inv.get("product_id"):
            product = await db.products.find_one({"id": inv.get("product_id")})

        if collection_product_ids is not None:
            if not product or product.get("id") not in collection_product_ids:
                continue

        variant = None
        if product:
            variant = next((v for v in product.get("variants", []) if v.get("id") == inv.get("variant_id")), None)

        location = await db.locations.find_one({"id": inv.get("location_id")})
        shelf = await db.shelves.find_one({"id": inv.get("shelf_id")}) if inv.get("shelf_id") else None

        product_title = product.get("title") if product else inv.get("product_title") or f"Unknown ({inv.get('variant_id')})"
        variant_title = (variant.get("title") if variant else None) or inv.get("variant_title")
        barcode = (variant.get("barcode") if variant else None) or inv.get("variant_barcode")
        upc_backup = (variant.get("upc_backup") if variant else None) or inv.get("variant_upc_backup")
        sku = (variant.get("sku") if variant else None) or inv.get("variant_sku")
        price = (variant.get("price") if variant else None) or inv.get("variant_price") or 0
        product_image = (product.get("image_base64") or product.get("image_url")) if product else inv.get("product_image")
        product_id = (product.get("id") if product else None) or inv.get("product_id") or inv.get("variant_id")

        if variant_ids is not None:
            if inv.get("variant_id") not in variant_ids:
                continue
        if product_ids is not None and (variant_ids is None):
            if product_id not in product_ids:
                continue
        if size_l:
            vt = (variant_title or "").lower()
            if size_l not in vt:
                continue
        if search_l:
            hay = f"{product_title} {variant_title or ''} {barcode or ''} {upc_backup or ''} {sku or ''}".lower()
            if search_l not in hay:
                continue

        rows.append({
            "product_id": product_id,
            "product_title": product_title,
            "variant_title": variant_title,
            "barcode": barcode,
            "upc_backup": upc_backup,
            "sku": sku,
            "price": price,
            "location_name": location.get("name") if location else "",
            "shelf_name": shelf.get("name") if shelf else "",
            "quantity": qty,
            "product_image": product_image
        })

    return rows
# ==================== EXPORT ROUTES ====================

@api_router.get("/export/excel")
async def export_excel(
    location_id: Optional[str] = None,
    collection_id: Optional[str] = None,
    search: Optional[str] = None,
    size: Optional[str] = None,
    product_ids: Optional[str] = None,
    variant_ids: Optional[str] = None,
    token: Optional[str] = None,
    current_user: dict = Depends(get_current_user_or_token)
):
    """Export inventory as Excel with optional filters"""
    # Get collection product IDs if filtered
    collection_product_ids = None
    if collection_id:
        collection = await db.collections.find_one({"id": collection_id})
        if collection:
            collection_product_ids = collection.get("product_ids", [])
    
    # Get inventory with product details (with Shopify fallback)
    rows = await _build_inventory_export_rows(
        location_id,
        collection_product_ids,
        search=search,
        size=size,
        product_ids=_parse_ids(product_ids),
        variant_ids=_parse_ids(variant_ids),
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Header style
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Product", "Variant", "Barcode", "SKU", "Price", "Location", "Shelf", "Quantity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for inv in rows:
        ws.cell(row=row, column=1, value=inv.get("product_title", "")).border = thin_border
        ws.cell(row=row, column=2, value=inv.get("variant_title") or "").border = thin_border
        ws.cell(row=row, column=3, value=inv.get("barcode") or "").border = thin_border
        ws.cell(row=row, column=4, value=inv.get("sku") or "").border = thin_border
        ws.cell(row=row, column=5, value=inv.get("price") or 0).border = thin_border
        ws.cell(row=row, column=6, value=inv.get("location_name") or "").border = thin_border
        ws.cell(row=row, column=7, value=inv.get("shelf_name") or "").border = thin_border
        ws.cell(row=row, column=8, value=inv.get("quantity", 0)).border = thin_border
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@api_router.get("/export/pdf")
async def export_pdf(
    location_id: Optional[str] = None,
    collection_id: Optional[str] = None,
    search: Optional[str] = None,
    size: Optional[str] = None,
    product_ids: Optional[str] = None,
    variant_ids: Optional[str] = None,
    token: Optional[str] = None,
    current_user: dict = Depends(get_current_user_or_token)
):
    """Export inventory as PDF catalog with optional filters"""
    # Get collection product IDs if filtered
    collection_product_ids = None
    if collection_id:
        collection = await db.collections.find_one({"id": collection_id})
        if collection:
            collection_product_ids = collection.get("product_ids", [])
    
    # Get inventory with product details (with Shopify fallback)
    rows = await _build_inventory_export_rows(
        location_id,
        collection_product_ids,
        search=search,
        size=size,
        product_ids=_parse_ids(product_ids),
        variant_ids=_parse_ids(variant_ids),
    )
    
    # Group by product
    products_data = {}
    for idx, inv in enumerate(rows):
        product_id = inv.get("product_id") or f"unknown-{idx}"
        if product_id not in products_data:
            products_data[product_id] = {
                "title": inv.get("product_title", ""),
                "image_base64": inv.get("product_image"),
                "variants": []
            }

        products_data[product_id]["variants"].append({
            "title": inv.get("variant_title", "") or "",
            "barcode": inv.get("barcode", "") or "",
            "price": inv.get("price", 0) or 0,
            "quantity": inv.get("quantity", 0),
            "location": inv.get("location_name", "") or "",
            "shelf": inv.get("shelf_name", "") or ""
        })
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1.6*cm, bottomMargin=1.6*cm)
    styles = getSampleStyleSheet()
    elements = []

    export_date = datetime.now().strftime('%d/%m/%Y')
    logo_path = (Path(__file__).resolve().parent / ".." / "frontend" / "LOGOSHARKDROP.png").resolve()

    def _draw_header_footer(canvas, doc):
        canvas.saveState()
        # Header
        header_y = A4[1] - 1.1 * cm
        if logo_path.exists():
            try:
                canvas.drawImage(str(logo_path), 1.2 * cm, A4[1] - 1.7 * cm, width=1.2 * cm, height=1.2 * cm, mask='auto')
            except Exception:
                pass
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(2.8 * cm, header_y, f"INVENTARIO SHARKDROP - {export_date}")

        # Footer
        canvas.setFont("Helvetica", 9)
        footer_text = "Proprietario: Giulio Fattore | FattoreH S.A.S | +39 342 756 9600"
        canvas.drawString(1.2 * cm, 0.9 * cm, footer_text)
        canvas.restoreState()

    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=28,
        spaceAfter=22
    )
    name_style = ParagraphStyle(
        'Name',
        parent=styles['Heading2'],
        fontSize=20,
        spaceAfter=8
    )
    sizes_style = ParagraphStyle(
        'Sizes',
        parent=styles['BodyText'],
        fontSize=15,
        leading=20
    )

    elements.append(Spacer(1, 12))

    def size_key(title: str) -> float:
        m = re.search(r"(\\d+(?:\\.\\d+)?)", (title or "").replace(",", "."))
        if not m:
            return 9999.0
        try:
            return float(m.group(1))
        except:
            return 9999.0

    for _, data in products_data.items():
        image_flowable = None
        img_src = data.get("image_base64")
        if img_src:
            try:
                img_bytes = None
                if isinstance(img_src, str) and img_src.startswith("data:image"):
                    img_bytes = base64.b64decode(img_src.split(",", 1)[1])
                elif isinstance(img_src, str) and img_src.startswith("http"):
                    async with aiohttp.ClientSession() as session:
                        async with session.get(img_src) as resp:
                            if resp.status == 200:
                                img_bytes = await resp.read()
                if img_bytes:
                    image_flowable = RLImage(BytesIO(img_bytes), width=3*cm, height=3*cm)
            except Exception:
                image_flowable = None

        variants = sorted(data["variants"], key=lambda v: size_key(v.get("title", "")))
        size_bits = []
        for v in variants:
            price = v.get("price", 0) or 0
            size_bits.append(f"{v.get('title','')}  EUR {price:.2f}")
        sizes_line = " | ".join(size_bits) if size_bits else ""

        text_block = Paragraph(
            f"<b>{data['title']}</b><br/>{sizes_line}",
            sizes_style
        )

        if image_flowable:
            row = [[image_flowable, text_block]]
            table = Table(row, colWidths=[3.2*cm, 13.3*cm])
        else:
            row = [[text_block]]
            table = Table(row, colWidths=[16.5*cm])

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('BOX', (0, 0), (-1, -1), 1, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.2*inch))

    doc.build(elements, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=catalog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )

@api_router.get("/export/csv")
async def export_csv(
    location_id: Optional[str] = None,
    collection_id: Optional[str] = None,
    search: Optional[str] = None,
    size: Optional[str] = None,
    product_ids: Optional[str] = None,
    variant_ids: Optional[str] = None,
    token: Optional[str] = None,
    current_user: dict = Depends(get_current_user_or_token)
):
    """Export inventory as CSV with optional filters"""
    collection_product_ids = None
    if collection_id:
        collection = await db.collections.find_one({"id": collection_id})
        if collection:
            collection_product_ids = collection.get("product_ids", [])

    rows = await _build_inventory_export_rows(
        location_id,
        collection_product_ids,
        search=search,
        size=size,
        product_ids=_parse_ids(product_ids),
        variant_ids=_parse_ids(variant_ids),
    )

    from io import StringIO
    string_io = StringIO()
    writer = csv.writer(string_io)
    writer.writerow(["Product", "Variant", "Barcode", "SKU", "Price", "Location", "Shelf", "Quantity"])
    for inv in rows:
        writer.writerow([
            inv.get("product_title", ""),
            inv.get("variant_title") or "",
            inv.get("barcode") or "",
            inv.get("sku") or "",
            inv.get("price") or 0,
            inv.get("location_name") or "",
            inv.get("shelf_name") or "",
            inv.get("quantity", 0)
        ])
    output = BytesIO(string_io.getvalue().encode("utf-8"))
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    """Get dashboard statistics"""
    # Total products
    total_products = await db.products.count_documents({"is_active": True})
    
    # Total inventory items
    pipeline = [
        {"$match": {"quantity": {"$gt": 0}}},
        {"$group": {"_id": None, "total": {"$sum": "$quantity"}}}
    ]
    inv_result = await db.inventory_levels.aggregate(pipeline).to_list(1)
    total_inventory = inv_result[0]["total"] if inv_result else 0
    
    # Locations count
    total_locations = await db.locations.count_documents({"is_active": True})
    
    # Today's transactions
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_transactions = await db.inventory_transactions.count_documents({
        "created_at": {"$gte": today_start}
    })
    
    # Recent transactions
    recent_tx = await db.inventory_transactions.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)

    # Inventory source breakdown
    source_breakdown = {}
    source_pipeline = [
        {"$group": {"_id": "$source", "total": {"$sum": "$quantity"}}}
    ]
    source_result = await db.inventory_levels.aggregate(source_pipeline).to_list(10)
    for row in source_result:
        source_breakdown[row.get("_id") or "unknown"] = row.get("total", 0)

    # Shopify sync status (for dashboard hints)
    sync_state = await db.shopify_sync_state.find_one({}, {"_id": 0})
    
    return {
        "total_products": total_products,
        "total_inventory": total_inventory,
        "total_locations": total_locations,
        "today_transactions": today_transactions,
        "recent_transactions": recent_tx,
        "inventory_sources": source_breakdown,
        "shopify_sync_status": sync_state.get("status") if sync_state else "never_synced",
        "shopify_last_sync_at": sync_state.get("last_sync_at") if sync_state else None
    }

# ==================== ANALYTICS & SALES ====================

def _parse_date_range(date: Optional[str] = None, month: Optional[str] = None, from_date: Optional[str] = None, to_date: Optional[str] = None):
    if date:
        start = datetime.strptime(date, "%Y-%m-%d")
        end = start + timedelta(days=1)
        return start, end
    if month:
        start = datetime.strptime(month + "-01", "%Y-%m-%d")
        if start.month == 12:
            end = datetime(start.year + 1, 1, 1)
        else:
            end = datetime(start.year, start.month + 1, 1)
        return start, end
    if from_date or to_date:
        start = datetime.strptime(from_date, "%Y-%m-%d") if from_date else datetime(1970, 1, 1)
        end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1) if to_date else datetime.utcnow() + timedelta(days=1)
        return start, end
    return None, None

@api_router.get("/analytics/summary")
async def get_analytics_summary(current_user: dict = Depends(get_current_user)):
    """Analytics summary for dashboard (today + weekly)"""
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today_start + timedelta(days=1)
    week_start = today_start - timedelta(days=6)

    # Today stats
    tx_today = await db.inventory_transactions.find(
        {"created_at": {"$gte": today_start, "$lt": tomorrow}},
        {"_id": 0}
    ).to_list(10000)
    received_today = sum(t.get("quantity", 0) for t in tx_today if t.get("transaction_type") == "receive")
    sold_today = sum(t.get("quantity", 0) for t in tx_today if t.get("transaction_type") == "sale")
    revenue_today = sum((t.get("sale_price") or 0) * (t.get("quantity") or 0) for t in tx_today if t.get("transaction_type") == "sale")

    # Weekly series (last 7 days)
    tx_week = await db.inventory_transactions.find(
        {"created_at": {"$gte": week_start, "$lt": tomorrow}},
        {"_id": 0}
    ).to_list(10000)
    series = []
    for i in range(7):
        day = week_start + timedelta(days=i)
        day_end = day + timedelta(days=1)
        day_tx = [t for t in tx_week if day <= t.get("created_at") < day_end]
        day_sales_qty = sum(t.get("quantity", 0) for t in day_tx if t.get("transaction_type") == "sale")
        day_sales_amount = sum((t.get("sale_price") or 0) * (t.get("quantity") or 0) for t in day_tx if t.get("transaction_type") == "sale")
        day_received = sum(t.get("quantity", 0) for t in day_tx if t.get("transaction_type") == "receive")
        series.append({
            "date": day.strftime("%Y-%m-%d"),
            "sales_qty": day_sales_qty,
            "sales_amount": day_sales_amount,
            "received_qty": day_received
        })

    return {
        "received_today": received_today,
        "sold_today": sold_today,
        "revenue_today": revenue_today,
        "weekly": series
    }

@api_router.get("/sales")
async def list_sales(
    date: Optional[str] = None,
    month: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """List sales with date filters"""
    start, end = _parse_date_range(date=date, month=month, from_date=from_date, to_date=to_date)
    query: Dict[str, Any] = {"transaction_type": "sale"}
    if start and end:
        query["created_at"] = {"$gte": start, "$lt": end}
    total = await db.inventory_transactions.count_documents(query)
    txs = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(offset).limit(limit).to_list(limit)

    variant_ids = [t.get("variant_id") for t in txs if t.get("variant_id")]
    products = await db.products.find({"variants.id": {"$in": variant_ids}}, {"_id": 0}).to_list(10000)
    variant_map = {}
    for p in products:
        for v in p.get("variants", []):
            variant_map[v.get("id")] = {
                "product_title": p.get("title"),
                "variant_title": v.get("title")
            }

    items = []
    for t in txs:
        vinfo = variant_map.get(t.get("variant_id"), {})
        items.append({
            "id": t.get("id"),
            "created_at": t.get("created_at"),
            "quantity": t.get("quantity", 0),
            "sale_price": t.get("sale_price", 0),
            "total_amount": (t.get("sale_price") or 0) * (t.get("quantity") or 0),
            "product_title": vinfo.get("product_title") or t.get("product_title"),
            "variant_title": vinfo.get("variant_title") or t.get("variant_title"),
        })
    return {"total": total, "sales": items}

@api_router.get("/sales/export/csv")
async def export_sales_csv(
    date: Optional[str] = None,
    month: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    token: Optional[str] = None,
    current_user: dict = Depends(get_current_user_or_token)
):
    """Export sales to CSV with date filters"""
    start, end = _parse_date_range(date=date, month=month, from_date=from_date, to_date=to_date)
    query: Dict[str, Any] = {"transaction_type": "sale"}
    if start and end:
        query["created_at"] = {"$gte": start, "$lt": end}
    txs = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)

    variant_ids = [t.get("variant_id") for t in txs if t.get("variant_id")]
    products = await db.products.find({"variants.id": {"$in": variant_ids}}, {"_id": 0}).to_list(10000)
    variant_map = {}
    for p in products:
        for v in p.get("variants", []):
            variant_map[v.get("id")] = {
                "product_title": p.get("title"),
                "variant_title": v.get("title")
            }

    from io import StringIO
    string_io = StringIO()
    writer = csv.writer(string_io)
    writer.writerow(["Product", "Variant", "Sale Date", "Quantity", "Sale Price", "Total Amount"])
    for t in txs:
        vinfo = variant_map.get(t.get("variant_id"), {})
        writer.writerow([
            vinfo.get("product_title") or t.get("product_title") or "",
            vinfo.get("variant_title") or t.get("variant_title") or "",
            t.get("created_at").strftime("%Y-%m-%d %H:%M") if t.get("created_at") else "",
            t.get("quantity", 0),
            t.get("sale_price", 0),
            (t.get("sale_price") or 0) * (t.get("quantity") or 0)
        ])
    output = BytesIO(string_io.getvalue().encode("utf-8"))
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )

# ==================== SETUP ====================

@api_router.get("/")
async def root():
    return {"message": "SharkDrop WMS API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[
        "http://localhost:19007",
        "http://127.0.0.1:19007",
        "http://192.168.0.16:19007",
        "http://localhost:8081",
        "http://127.0.0.1:8081",
        "http://192.168.0.16:8081",
        "https://gestionale.sharkdrop.it",
    ],
    allow_origin_regex=r"^http://(localhost|127\.0\.0\.1|192\.168\.0\.16)(:\d+)?$",
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_db_client():
    # Create indexes
    await db.users.create_index("username", unique=True)
    await db.users.create_index("id", unique=True)
    await db.products.create_index("shopify_product_id")
    await db.products.create_index("id")
    await db.products.create_index("is_active")
    await db.products.create_index("variants.barcode")
    await db.products.create_index("variants.upc_backup")
    await db.shelves.create_index("barcode", unique=True)
    await db.inventory_levels.create_index([("variant_id", 1), ("location_id", 1), ("shelf_id", 1)])
    await db.action_logs.create_index("created_at")
    await db.action_logs.create_index("user_id")
    await db.action_logs.create_index("action_type")
    await db.inventory_transactions.create_index("user_id")
    await db.inventory_transactions.create_index("created_at")
    await db.system_logs.create_index("created_at")
    await db.stall_items.create_index("status")
    await db.stall_items.create_index("created_at")
    await db.stockx_lookups.create_index("created_at")
    await db.stockx_lookups.create_index("barcode")
    await db.stockx_import_jobs.create_index("created_at")
    await db.stockx_import_jobs.create_index("status")
    
    # Create or update default admin
    admin = await db.users.find_one({"username": "admin"})
    if not admin:
        admin_user = User(
            username="admin",
            password_hash=get_password_hash("SharkAdmin!"),
            role=UserRole.ADMIN
        )
        await db.users.insert_one(admin_user.dict())
        logger.info("Default admin user created: admin/SharkAdmin!")
    else:
        # Update admin password if still using old default
        if verify_password("admin123", admin.get("password_hash", "")):
            await db.users.update_one(
                {"username": "admin"},
                {"$set": {"password_hash": get_password_hash("SharkAdmin!")}}
            )
            logger.info("Admin password updated to SharkAdmin!")

    # Env sanity checks (non-blocking)
    shop_domain, access_token, _ = get_shopify_config()
    if not shop_domain or not access_token:
        await log_system_event("warning", "Shopify credentials missing or incomplete")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


